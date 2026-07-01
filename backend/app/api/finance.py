"""Finance API - supplier reconciliation, user statistics."""

from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, FileResponse

from app.services import finance_service

router = APIRouter(prefix="/api/finance", tags=["finance"])


class FinanceQueryError(Exception):
    pass


# ── Common helpers ──

@router.get("/log-tables")
async def log_tables(site: str = Query(...)):
    return {"tables": await finance_service.get_log_tables(site)}


@router.get("/usernames")
async def usernames(
    site: str = Query(...),
    table: str = Query(""),
):
    return {"usernames": await finance_service.get_usernames(site, table)}


@router.get("/table-dates")
async def table_dates(table: str = Query(...)):
    start, end = finance_service.parse_log_table_dates(table)
    return {"start": start, "end": end}


# ── Supplier reconciliation ──

@router.get("/supplier")
async def supplier_query(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    supplier_name: str = Query(""),
):
    try:
        rows = await finance_service.supplier_query(site, table, username, date_start, date_end, supplier_name)
        return {"rows": rows}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/supplier/export")
async def supplier_export(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    supplier_name: str = Query(""),
):
    rows = await finance_service.supplier_query(site, table, username, date_start, date_end, supplier_name)
    if not rows:
        raise HTTPException(400, detail="无数据可导出")
    content = _build_supplier_excel(rows)
    ds = date_start.replace("-", "")
    de = date_end.replace("-", "")
    filename = f"supplier{ds}_{de}_{username}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_supplier_excel(rows):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商对账"

    headers = list(rows[0].keys()) + ["折扣", "实付金额（USD）"]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    total_col = headers.index("总费用（USD）") + 1
    for r in rows:
        ws.append(list(r.values()) + ["", ""])

    ws.append(["" for _ in headers])
    total_row = ["" for _ in headers]
    total_row[0] = "合计"
    total_row[total_col - 1] = (
        f"=SUM({openpyxl.utils.get_column_letter(total_col)}2"
        f":{openpyxl.utils.get_column_letter(total_col)}{len(rows)+1})"
    )
    ws.append(total_row)

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── User statistics ──

@router.get("/user-stats")
async def user_stats(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    granularity: str = Query(""),
):
    try:
        glist = [x.strip() for x in granularity.split(",") if x.strip()] if granularity else []
        show_model = "model" in glist
        show_token = "token" in glist
        monthly = await finance_service.user_monthly(site, table, username, date_start, date_end, show_model)
        daily = await finance_service.user_daily(site, table, username, date_start, date_end, show_model, show_token)
        return {"monthly": monthly, "daily": daily}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/user-stats/detail")
async def user_stats_detail(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    page: int = Query(1),
    size: int = Query(20),
):
    try:
        return await finance_service.user_detail(site, table, username, date_start, date_end, page, size)
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/user-stats/export")
async def user_stats_export(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    with_platform: str = Query("0"),
):
    monthly = await finance_service.user_monthly(site, table, username, date_start, date_end)
    daily = await finance_service.user_daily(site, table, username, date_start, date_end)
    detail_res = await finance_service.user_detail(site, table, username, date_start, date_end, page=1, page_size=1000000)
    detail = detail_res.get("data", [])
    if not monthly and not daily and not detail:
        raise HTTPException(400, detail="无数据可导出")
    content = _build_user_stats_excel(monthly, daily, detail, with_platform == "1")
    ds = date_start.replace("-", "")
    de = date_end.replace("-", "")
    filename = f"{username}_{ds}_{de}.xlsx"
    encoded = quote(filename)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/user-stats/export-async")
async def user_stats_export_async(body: dict):
    site = body.get("site", "")
    table = body.get("table", "")
    username = body.get("username", "")
    date_start = body.get("date_start", "")
    date_end = body.get("date_end", "")
    with_platform = body.get("with_platform", False)
    with_detail = body.get("with_detail", True)
    with_total_cost = body.get("with_total_cost", True)
    granularity = body.get("granularity", "")
    glist = [x.strip() for x in granularity.split(",") if x.strip()] if granularity else []
    show_model = "model" in glist
    show_token = "token" in glist
    if not site or not table:
        raise HTTPException(400, detail="site, table 不能为空")
    task_id = finance_service.start_export_task(
        site, table, username, date_start, date_end, with_platform, with_detail, show_model, show_token, with_total_cost
    )
    return {"task_id": task_id}


@router.get("/user-stats/export-status")
async def user_stats_export_status(task_id: str = Query(...)):
    status = finance_service.get_export_status(task_id)
    if not status:
        raise HTTPException(404, detail="任务不存在")
    return status


@router.get("/user-stats/export-download")
async def user_stats_export_download(task_id: str = Query(...)):
    import os
    status = finance_service.get_export_status(task_id)
    if not status:
        raise HTTPException(404, detail="任务不存在")
    if status["status"] != "done":
        raise HTTPException(400, detail="文件尚未生成完毕")
    file_path = status.get("file_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(500, detail="文件不存在")
    file_size = os.path.getsize(file_path)
    if file_size == 0:
        raise HTTPException(500, detail="生成的文件为空，请重试")
    username = status.get("username", "export")
    ds = status.get("date_start", "").replace("-", "")
    de = status.get("date_end", "").replace("-", "")
    filename = f"{username}_{ds}_{de}.xlsx" if ds else f"{username}.xlsx"
    encoded = quote(filename)
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
    """Remove specified columns from each row dict."""
    if not rows:
        return rows
    return [{k: v for k, v in r.items() if k not in strip_keys} for r in rows]


def _append_sheet_with_totals(ws, rows, total_fields):
    """Write data rows and append a SUM total row for specified fields."""
    if not rows:
        return
    import openpyxl
    from openpyxl.styles import Font, Alignment
    headers = list(rows[0].keys())
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(list(r.values()))

    # Build total row with SUM formulas at correct column positions
    last_data_row = len(rows) + 1
    total_row = ["合计"] + ["" for _ in range(len(headers) - 1)]
    for field in total_fields:
        if field in headers:
            ci = headers.index(field) + 1
            cl = openpyxl.utils.get_column_letter(ci)
            total_row[ci - 1] = f"=SUM({cl}2:{cl}{last_data_row})"
    ws.append([])
    ws.append(total_row)


def _build_user_stats_excel(monthly, daily, detail=None, with_platform=False):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    _PLATFORM_KEYS = {"平台额度"}

    if not with_platform:
        monthly = _strip_platform(monthly, _PLATFORM_KEYS)
        daily = _strip_platform(daily, _PLATFORM_KEYS)
        detail = _strip_platform(detail, _PLATFORM_KEYS)

    # Fields that should get SUM totals — numeric data columns, not IDs/names/unit-prices
    _NO_TOTAL_KEYS = {"结算周期", "用户名", "模型名", "用户ID", "Token名称", "日期", "时间", "序号",
                       "输入单价", "输出单价", "读取缓存单价", "创建缓存5M单价", "创建缓存1H单价"}

    def _total_fields(rows):
        if not rows:
            return []
        return [k for k in rows[0].keys() if k not in _NO_TOTAL_KEYS]

    wb = openpyxl.Workbook()
    ws_month = wb.active
    ws_month.title = "月汇总"
    if monthly:
        _append_sheet_with_totals(ws_month, monthly, _total_fields(monthly))

    ws_daily = wb.create_sheet("日统计")
    if daily:
        _append_sheet_with_totals(ws_daily, daily, _total_fields(daily))

    if detail:
        ws_detail = wb.create_sheet("用户明细")
        _append_sheet_with_totals(ws_detail, detail, _total_fields(detail))

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Site monthly report ──

@router.get("/site-report/preview")
async def site_report_preview(
    site: str = Query(...),
    table: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
):
    return await finance_service.site_report_preview(site, table, date_start, date_end)


@router.post("/site-report/generate")
async def site_report_generate(body: dict):
    site = body.get("site")
    table = body.get("table")
    date_start = body.get("date_start")
    date_end = body.get("date_end")
    output_root = body.get("output_root", "E:/Workspaces/claude/billsum/out")

    if not all([site, table, date_start, date_end]):
        raise HTTPException(400, detail="参数不完整")

    return await finance_service.generate_all_reports(
        site, table, date_start, date_end, output_root,
    )


@router.post("/site-report/generate-zip")
async def site_report_generate_zip(body: dict):
    site = body.get("site")
    table = body.get("table")
    date_start = body.get("date_start")
    date_end = body.get("date_end")

    if not all([site, table, date_start, date_end]):
        raise HTTPException(400, detail="参数不完整")

    zip_bytes = await finance_service.generate_reports_zip(site, table, date_start, date_end)
    ym = date_start.replace("-", "")[:6]
    filename = f"{site}_report{ym}.zip"
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
