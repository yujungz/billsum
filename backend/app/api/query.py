"""Query API - list, query, delete and export tables."""

import csv
import io
import os
import re
import subprocess
import tempfile

from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from fastapi.responses import Response

from app.config import AppConfig
from app.services import query_service, parser_service

router = APIRouter(prefix="/api/query", tags=["query"])

TABLE_NAME_RE = re.compile(r'^[a-zA-Z_][a-zA-Z0-9_]*$')


def _validate_table(name: str):
    if not TABLE_NAME_RE.match(name):
        raise HTTPException(400, detail=f"Invalid table name: {name}")


@router.get("/tables")
async def list_tables(site: str = Query(...), type: str = Query("raw")):
    if type == "raw":
        return {"tables": await query_service.list_raw_tables(site)}
    else:
        return {"tables": await query_service.list_output_tables(site)}


@router.get("/log-tables")
async def list_log_tables(site: str = Query(...)):
    return {"tables": await query_service.list_log_tables(site)}


@router.get("/columns")
async def get_columns(site: str = Query(...), table: str = Query(...)):
    _validate_table(table)
    return {"columns": await query_service.get_table_columns(site, table)}


@router.get("/data")
async def query_data(
    site: str = Query(...),
    table: str = Query(...),
    page: int = Query(1),
    size: int = Query(20),
    filters: str = Query(None),
    time_order: str = Query("desc"),
):
    _validate_table(table)
    import json
    f = json.loads(filters) if filters else None
    return await query_service.query_table(site, table, page, size, f, time_order)


@router.delete("/table")
async def delete_table(site: str = Query(...), table: str = Query(...)):
    _validate_table(table)
    await query_service.delete_table(site, table)
    return {"success": True, "message": f"Table {table} deleted"}


@router.get("/export")
async def export_table(
    site: str = Query(...),
    table: str = Query(...),
    format: str = Query("xlsx"),
    filters: str = Query(None),
    fields: str = Query(None),
):
    _validate_table(table)
    import json
    f = json.loads(filters) if filters else None
    fn = f"{site}_{table}"

    if format == "sql":
        content = query_service.export_table_sql(site, table)
        return Response(
            content=content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fn}.sql"},
        )

    columns, rows = await query_service.export_all_data(site, table, f)
    cols = _apply_fields(columns, fields)

    if format == "csv":
        output = io.StringIO()
        output.write('﻿')
        writer = csv.writer(output)
        writer.writerow([c['label'] for c in cols])
        for row in rows:
            writer.writerow([str(row.get(c['name'], '')) for c in cols])
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fn}.csv"},
        )

    # xlsx - openpyxl
    content = _build_xlsx(cols, rows)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}.xlsx"},
    )


def _apply_fields(columns, fields_json):
    """Filter/rename columns per the field-selection config.
    fields_json: JSON '[{"name":..,"label":..}, ...]' (selected fields in order).
    None/invalid → all columns, label = field name.
    """
    default = [{"name": c["name"], "label": c["name"]} for c in columns]
    if not fields_json:
        return default
    import json
    try:
        flds = json.loads(fields_json)
    except Exception:
        return default
    db_names = {c["name"] for c in columns}
    out = []
    for x in flds:
        nm = x.get("name")
        if nm in db_names:
            out.append({"name": nm, "label": x.get("label") or nm})
    return out or default


def _build_xlsx(columns, rows):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    bold = Font(bold=True)
    for ci, c in enumerate(columns, 1):
        ws.cell(row=1, column=ci, value=c['label']).font = bold
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(str(c['label'])) * 2, 12)
    for ri, row in enumerate(rows, 2):
        for ci, c in enumerate(columns, 1):
            ws.cell(row=ri, column=ci, value=row.get(c['name'], ''))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


COLUMN_ALIASES = {
    # channels / ex_channels
    "渠道ID": "id", "渠道名称": "name", "采购员": "buyer", "供应商": "supplier",
    "原始折扣": "discount_orig", "折扣": "discount", "原始折扣 折/刀": "discount_orig",
    # tokens / ex_tokens
    "TokenID": "id", "UserID": "user_id",
    # users / ex_users
    "用户ID": "id", "用户名称": "name", "用户名": "name",
    "备注": "remark", "销售人员": "seller", "销售员": "seller",
}


def _resolve_columns(raw_headers: list[str], db_columns: set[str]) -> list[str | None]:
    """Map file headers to db columns. Stops at first empty header to ignore trailing cols."""
    result: list[str | None] = []
    for h in raw_headers:
        h = h.strip()
        if not h:
            break  # stop at first empty column
        if h in db_columns:
            result.append(h)
            continue
        mapped = COLUMN_ALIASES.get(h)
        if mapped and mapped in db_columns:
            result.append(mapped)
        else:
            result.append(None)
    return result


def _map_row(row: tuple, col_map: list[str | None]) -> list | None:
    """Extract values from a row, only keeping columns that mapped to db fields."""
    vals = []
    for i, col in enumerate(col_map):
        if col is None:
            continue
        v = row[i] if i < len(row) else None
        vals.append(None if v is None else v)
    return vals if vals else None


def _map_row_csv(row: list[str], col_map: list[str | None]) -> list | None:
    """Extract values from a CSV row, only keeping columns that mapped to db fields."""
    vals = []
    for i, col in enumerate(col_map):
        if col is None:
            continue
        v = row[i] if i < len(row) else ""
        vals.append(None if v == "" else v)
    return vals if vals else None


@router.post("/import")
async def import_sql(site: str = Query(...), table: str = Query(...), file: UploadFile = File(...)):
    _validate_table(table)
    config = AppConfig.load()
    db_name = config.db_name(site)

    filename = file.filename or "import.sql"
    ext = os.path.splitext(filename)[1].lower()

    content = await file.read()

    if ext == ".sql":
        mc = config.mysql
        suffix = ext
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False, mode="wb") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            cmd = [
                "mysql",
                f"--host={mc.host}", f"--port={mc.port}",
                f"--user={mc.user}", f"--password={mc.password}",
                "--default-character-set=utf8mb4", "--skip-ssl",
                db_name,
            ]
            with open(tmp_path, "r", encoding="utf-8") as f:
                proc = subprocess.run(cmd, stdin=f, capture_output=True, text=True, timeout=600)
            if proc.returncode != 0:
                raise HTTPException(400, detail=f"Import failed: {proc.stderr[:500]}")
        finally:
            os.unlink(tmp_path)

    elif ext in (".xlsx", ".xls"):
        import openpyxl
        from app import database as db

        await db.execute(f"TRUNCATE TABLE `{table}`", db=db_name)

        # get actual db column names
        db_cols = await query_service.get_table_columns(site, table)
        db_col_set = {c["name"] for c in db_cols}

        buf = io.BytesIO(content)
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        col_map = _resolve_columns(raw_headers, db_col_set)  # list[str|None], same length as raw_headers
        mapped_cols = [c for c in col_map if c is not None]
        if not mapped_cols:
            raise HTTPException(400, detail="文件列名与表字段无法对应")

        cols_sql = ", ".join(f"`{c}`" for c in mapped_cols)
        placeholders = ", ".join(["%s"] * len(mapped_cols))
        sql = f"INSERT INTO `{table}` ({cols_sql}) VALUES ({placeholders})"

        cols_sql = ", ".join(f"`{c}`" for c in col_map)
        placeholders = ", ".join(["%s"] * len(col_map))
        sql = f"INSERT INTO `{table}` ({cols_sql}) VALUES ({placeholders})"

        count = 0
        batch = []
        for row in rows_iter:
            vals = _map_row(row, col_map)
            if vals is not None:
                batch.append(vals)
                if len(batch) >= 500:
                    await db.execute_many(sql, batch, db=db_name)
                    count += len(batch)
                    batch = []
        if batch:
            await db.execute_many(sql, batch, db=db_name)
            count += len(batch)
        wb.close()

    elif ext == ".csv":
        import csv as csv_mod
        from app import database as db

        await db.execute(f"TRUNCATE TABLE `{table}`", db=db_name)

        db_cols = await query_service.get_table_columns(site, table)
        db_col_set = {c["name"] for c in db_cols}

        text = content.decode("utf-8-sig")
        reader = csv_mod.reader(io.StringIO(text))
        raw_headers = [h.strip() for h in next(reader)]
        col_map = _resolve_columns(raw_headers, db_col_set)
        mapped_cols = [c for c in col_map if c is not None]
        if not mapped_cols:
            raise HTTPException(400, detail="文件列名与表字段无法对应")

        cols_sql = ", ".join(f"`{c}`" for c in mapped_cols)
        placeholders = ", ".join(["%s"] * len(mapped_cols))
        sql = f"INSERT INTO `{table}` ({cols_sql}) VALUES ({placeholders})"

        count = 0
        batch = []
        for row in reader:
            vals = _map_row_csv(row, col_map)
            if vals is not None:
                batch.append(vals)
                if len(batch) >= 500:
                    await db.execute_many(sql, batch, db=db_name)
                    count += len(batch)
                    batch = []
        if batch:
            await db.execute_many(sql, batch, db=db_name)
            count += len(batch)

    else:
        raise HTTPException(400, detail=f"不支持的文件格式: {ext}")

    return {"success": True, "message": "Import completed"}


@router.post("/parse")
async def parse_table(site: str = Query(...), table: str = Query(...)):
    if table not in parser_service.PARSEABLE_TABLES:
        raise HTTPException(400, detail=f"Table '{table}' is not parseable. Use: channels, tokens, users")

    parse_fn = {
        'channels': parser_service.parse_channels,
        'tokens': parser_service.parse_tokens,
        'users': parser_service.parse_users,
    }[table]

    result = await parse_fn(site)

    content = parser_service.build_excel_bytes(table, result["excel_headers"], result["excel_data"])

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={site}_ex_{table}.xlsx"},
    )
