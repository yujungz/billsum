"""Statistics API."""

import io
import json
import logging
import os
import time
import uuid
from pathlib import Path

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import Response, FileResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font

from app import database as db
from app.config import AppConfig
from app.services import stats_service

log = logging.getLogger(__name__)

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/stats", tags=["stats"])

# ── helpers shared by the original stats + the new detail sheet ──

FIELD_SQL = {
    "input_tokens": "l.prompt_tokens",
    "input_unit_price": "ROUND(l.model_ratio*2, 6)",
    "input_cost": "ROUND(l.group_ratio*l.model_ratio*2*l.prompt_tokens/1000000, 6)",
    "output_tokens": "l.completion_tokens",
    "output_unit_price": "ROUND(l.model_ratio*2*l.completion_ratio, 6)",
    "output_cost": "ROUND(l.group_ratio*l.model_ratio*2*l.completion_ratio*l.completion_tokens/1000000, 6)",
    "cache_read_tokens": "l.cache_tokens",
    "cache_read_unit_price": "ROUND(l.model_ratio*2*l.cache_ratio, 6)",
    "cache_read_cost": "ROUND(l.group_ratio*l.model_ratio*2*l.cache_ratio*l.cache_tokens/1000000, 6)",
    "cache_create_5m_tokens": "l.cache_creation_tokens_5m",
    "cache_create_5m_unit_price": "ROUND(l.model_ratio*2*1.25, 6)",
    "cache_create_5m_cost": "ROUND(l.group_ratio*l.model_ratio*2*1.25*l.cache_creation_tokens_5m/1000000, 6)",
    "cache_create_1h_tokens": "GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m)",
    "cache_create_1h_unit_price": "ROUND(l.model_ratio*2*2.00, 6)",
    "cache_create_1h_cost": "ROUND(l.group_ratio*l.model_ratio*2*2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m)/1000000, 6)",
    "cache_create_tokens": "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m)",
    "cache_create_cost": "ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m + 2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m))/1000000, 6)",
    "cache_total_tokens": "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m) + l.cache_tokens",
    "cache_total_cost": "ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m + l.cache_ratio*l.cache_tokens + 2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m))/1000000, 6)",
    "total_tokens": "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m) + l.cache_tokens + l.completion_tokens + l.prompt_tokens",
    "total_cost": "ROUND(l.group_ratio*l.model_ratio*2*(l.prompt_tokens + l.completion_ratio*l.completion_tokens + l.cache_ratio*l.cache_tokens + 1.25*l.cache_creation_tokens_5m + 2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m))/1000000, 6)",
    "platform_quota": "l.quota*2/1000000",
}


def _col_width(label: str) -> float:
    w = len(label) * 1.8 + 2
    return max(w, 8)


def _detail_where(filters: dict | None):
    """Build WHERE clause for the detail SQL (no windup_type filter needed)."""
    conditions = ["l.windup_type < 2"]
    params = []
    if filters:
        ds = filters.get("date_start")
        de = filters.get("date_end")
        if ds:
            conditions.append("l.created_at+28800 >= UNIX_TIMESTAMP(%s)")
            params.append(f"{ds} 00:00:00")
        if de:
            conditions.append("l.created_at+28800 <= UNIX_TIMESTAMP(%s)")
            params.append(f"{de} 23:59:59")
    return " AND ".join(conditions), params


def _build_detail_columns(show_channel_name: bool, fields_json: str) -> list[dict]:
    """Build column definitions for the detail sheet.
    Returns list of {key, label} in display order."""
    cols = []
    # fixed: id, created_at, created_date
    cols.append({"key": "id", "label": "id"})
    cols.append({"key": "created_at", "label": "时间戳"})
    cols.append({"key": "created_date", "label": "创建日期"})
    # user fields
    cols.append({"key": "user_id", "label": "用户ID"})
    cols.append({"key": "username", "label": "用户"})
    cols.append({"key": "channel_id", "label": "渠道标识"})
    if show_channel_name:
        cols.append({"key": "channel_name", "label": "渠道名称"})
    cols.append({"key": "model_name", "label": "模型名称"})
    cols.append({"key": "token_id", "label": "Token标识"})
    cols.append({"key": "token_name", "label": "Token名称"})
    cols.append({"key": "group", "label": "分组"})
    cols.append({"key": "call_count", "label": "调用次数"})
    # dynamic fields from 字段选择
    if fields_json:
        try:
            flds = json.loads(fields_json)
            for x in flds:
                nm = x.get("name")
                if nm in FIELD_SQL:
                    cols.append({"key": nm, "label": x.get("label") or nm})
        except Exception:
            log.warning("failed to parse detail fields JSON", exc_info=True)
    return cols


def _build_detail_sql(table: str, detail_cols: list[dict], show_channel_name: bool,
                       where: str) -> str:
    """Build SELECT SQL for detail query (no ORDER BY/pagination; caller adds them)."""
    selects = []
    for c in detail_cols:
        k = c["key"]
        if k == "channel_name" and not show_channel_name:
            continue
        if k in FIELD_SQL:
            selects.append(f"{FIELD_SQL[k]} AS `{c['label']}`")
        elif k in _DETAIL_FIXED_SQL:
            selects.append(f"{_DETAIL_FIXED_SQL[k]} AS `{c['label']}`")
    return f"SELECT {', '.join(selects)} FROM `{table}` l WHERE {where}"


_DETAIL_FIXED_SQL = {
    "id": "l.id",
    "created_at": "l.created_at",
    "created_date": "FROM_UNIXTIME(l.created_at+28800)",
    "user_id": "l.user_id",
    "username": "l.username",
    "channel_id": "l.channel_id",
    "channel_name": "l.channel_name",
    "model_name": "l.model_name",
    "token_id": "l.token_id",
    "token_name": "l.token_name",
    "group": "l.`group`",
    "call_count": "1",
}

_CHUNK_SIZE = 50000


async def _write_detail_sheets(wb, config, detail_cols, where, params):
    """Streaming detail export with keyset pagination (no OFFSET).
    Uses `l.id < ?` which leverages the PRIMARY KEY index, O(n) regardless of depth."""
    from openpyxl.styles import Font
    app_config = AppConfig.load()
    db_name = config.db_name
    table = config.table_name

    # count (approximate; the loop stops naturally when no more rows)
    count_sql = f"SELECT COUNT(*) AS total FROM `{table}` l WHERE {where}"
    row = await db.fetch_one(count_sql, params, db=db_name)
    total = row["total"] if row else 0
    if not total:
        return

    headers = [c["label"] for c in detail_cols]

    # Build detail SQL once (no ORDER BY/pagination; keyset pagination added per batch)
    show_ch_name = any(c["key"] == "channel_name" for c in detail_cols)
    base_sql = _build_detail_sql(table, detail_cols, show_ch_name, where)
    LIMIT = 50000

    last_id = None
    sheet_idx = 0
    ws = None
    rows_in_sheet = 0
    processed = 0

    while processed < total:
        if last_id is None:
            sql = f"{base_sql} ORDER BY l.id DESC LIMIT {LIMIT}"
            chunk = await db.fetch_all(sql, params, db=db_name)
        else:
            sql = f"{base_sql} AND l.id < %s ORDER BY l.id DESC LIMIT {LIMIT}"
            chunk = await db.fetch_all(sql, params + [last_id], db=db_name)
        if not chunk:
            break

        for row_dict in chunk:
            if ws is None or rows_in_sheet >= 1000000:
                sheet_idx += 1
                name = "日志明细" if sheet_idx == 1 else f"日志明细_{sheet_idx}"
                ws = wb.create_sheet(name)
                rows_in_sheet = 0
                ws.append(headers)

            vals = []
            for cdef in detail_cols:
                v = row_dict.get(cdef["label"])
                vals.append(float(v) if v is not None and hasattr(v, "__float__") else v)
            ws.append(vals)
            rows_in_sheet += 1

        # Keyset: last row's id becomes the bound for the next batch
        last_id = chunk[-1].get("id")
        processed += len(chunk)

    log.info("detail export %s / %s total=%d written=%d", db_name, table, total, processed)


# ── Async detail export task machinery ──

DATA_DIR = Path(os.getenv("DATA_DIR", "/app/data"))
_STATS_EXPORT_TASKS: dict[str, dict] = {}


def start_stats_detail_task(site: str, table_name: str, filters: dict | None,
                            show_channel_name: bool, fields: str) -> str:
    """Start a background detail export task, return task_id immediately."""
    task_id = uuid.uuid4().hex[:8]
    _STATS_EXPORT_TASKS[task_id] = {
        "task_id": task_id,
        "status": "running",
        "progress": "排队中",
        "start_time": time.time(),
        "end_time": None,
        "file_path": None,
        "error": None,
        "site": site,
        "table_name": table_name,
    }

    async def _run():
        try:
            await _run_stats_detail(task_id, site, table_name, filters, show_channel_name, fields)
            _STATS_EXPORT_TASKS[task_id]["status"] = "done"
        except Exception as e:
            log.exception("stats detail export task %s failed", task_id)
            _STATS_EXPORT_TASKS[task_id]["status"] = "failed"
            _STATS_EXPORT_TASKS[task_id]["error"] = str(e)
        finally:
            _STATS_EXPORT_TASKS[task_id]["end_time"] = time.time()

    import asyncio
    asyncio.create_task(_run())
    return task_id


async def _run_stats_detail(task_id: str, site: str, table_name: str,
                            filters: dict | None,
                            show_channel_name: bool, fields: str):
    task = _STATS_EXPORT_TASKS[task_id]
    task["progress"] = "查询列配置..."
    show_ch = show_channel_name
    detail_cols = _build_detail_columns(show_ch, fields)
    if not detail_cols:
        task["error"] = "无可导出列"
        return

    where, params = _detail_where(filters)
    config_req = type("Config", (), {"db_name": f"sum_{site}", "table_name": table_name})()

    task["progress"] = "准备写入..."
    wb = openpyxl.Workbook(write_only=True)
    await _write_detail_sheets(wb, config_req, detail_cols, where, params)

    # Save to temp file
    out_dir = DATA_DIR / "stats_detail" / task_id
    out_dir.mkdir(parents=True, exist_ok=True)
    file_path = out_dir / "detail.xlsx"
    wb.save(str(file_path))

    task["file_path"] = str(file_path)
    task["progress"] = "完成"


def get_stats_detail_status(task_id: str) -> dict | None:
    t = _STATS_EXPORT_TASKS.get(task_id)
    if not t:
        return None
    elapsed = (t["end_time"] or time.time()) - t["start_time"]
    return {
        "task_id": task_id,
        "status": t["status"],
        "progress": t["progress"],
        "error": t.get("error"),
        "elapsed": round(elapsed, 1),
    }


def get_stats_detail_download(task_id: str) -> tuple[str, str] | None:
    t = _STATS_EXPORT_TASKS.get(task_id)
    if not t or t["status"] not in ("done", "failed"):
        return None
    path = t.get("file_path")
    if not path or not os.path.exists(path):
        return None
    fn = f"{t.get('site', 'stats')}_{t.get('table_name', 'detail')}_明细.xlsx"
    return path, fn


# ── Pydantic model ──


class StatsRequest(BaseModel):
    site: str
    table_name: str
    group_by: list[str] = []
    filters: dict | None = None
    show_zero: bool = True
    show_channel_name: bool = False
    show_log_detail: bool = False
    fields: str = ""


# ── Endpoints ──


@router.post("/query")
async def query_stats(req: StatsRequest):
    result = await stats_service.query_stats(
        req.site, req.table_name, req.group_by, req.filters, req.show_zero,
        show_channel_name=req.show_channel_name,
    )
    return {"data": result}


@router.get("/distinct")
async def get_distinct(site: str = Query(...), table: str = Query(...), field: str = Query(...)):
    values = await stats_service.get_distinct_values(site, table, field)
    return {"values": values}


@router.post("/export")
async def export_stats(req: StatsRequest):
    result = await stats_service.query_stats(
        req.site, req.table_name, req.group_by, req.filters, req.show_zero,
        show_channel_name=req.show_channel_name,
    )
    if not result:
        return Response(content=b"", status_code=204)

    # ─── stats sheet ───
    col_def = [
        ("period_month", "月份"),
        ("period_day", "日期"),
        ("user_id", "用户ID"),
        ("username", "用户名"),
        ("channel_id", "渠道ID"),
        ("channel_name", "渠道"),
        ("model_name", "模型"),
        ("group_name", "分组"),
        ("token_name", "Token名称"),
        ("cn_buyer1", "采购员"),
        ("cn_supplier1", "供应商"),
        ("us_salesperson", "销售员"),
        ("call_count", "调用次数"),
        ("input_tokens", "输入token"),
        ("input_unit_price", "输入单价"),
        ("input_cost", "输入费用"),
        ("output_tokens", "输出token"),
        ("output_unit_price", "输出单价"),
        ("output_cost", "输出费用"),
        ("cache_read_tokens", "读取缓存token"),
        ("cache_read_unit_price", "读取缓存单价"),
        ("cache_read_cost", "读取缓存费用"),
        ("cache_create_5m_tokens", "创建缓存5M-token"),
        ("cache_create_5m_unit_price", "创建缓存5M单价"),
        ("cache_create_5m_cost", "创建缓存5M费用"),
        ("cache_create_1h_tokens", "创建缓存1H-token"),
        ("cache_create_1h_unit_price", "创建缓存1H单价"),
        ("cache_create_1h_cost", "创建缓存1H费用"),
        ("cache_create_tokens", "创建缓存token"),
        ("cache_create_cost", "创建缓存费用"),
        ("cache_total_tokens", "缓存总token"),
        ("cache_total_cost", "缓存总费用"),
        ("total_tokens", "总消耗token"),
        ("total_cost", "消费额度"),
        ("platform_quota", "平台额度"),
    ]
    visible = [(k, l) for k, l in col_def if any(r.get(k) is not None for r in result)]

    # 字段选择过滤：若前端传了 fields，只保留被选中的列（按前端传入顺序）
    if req.fields:
        try:
            flds = {x["name"] for x in json.loads(req.fields)}
            # col_def 的非可计算字段（如粒度列）始终保留，不在 fields 中但要保留
            # 粒度列属于固定列，不在 FIELD_SQL 中，保留
            data_keys = set(k for k, _ in col_def if k not in FIELD_SQL)
            visible = [(k, l) for k, l in visible if k in data_keys or k in flds]
        except Exception:
            pass

    if req.group_by:
        date_col = "period_day" if "day" in req.group_by else "period_month"
        result.sort(key=lambda r: (r.get(date_col) or ""), reverse=True)

    no_sum = {"period_month", "period_day", "user_id", "username",
              "channel_id", "channel_name", "model_name", "group_name",
              "token_name", "cn_buyer1", "cn_supplier1", "us_salesperson"}
    sum_cols = [(ik, k) for ik, (k, l) in enumerate(visible) if k not in no_sum]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stats"
    bold = Font(bold=True)
    for ci, (_, label) in enumerate(visible, 1):
        ws.cell(row=1, column=ci, value=label).font = bold
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = _col_width(label)
    for ri, row in enumerate(result, 2):
        for ci, (key, _) in enumerate(visible, 1):
            v = row.get(key)
            ws.cell(row=ri, column=ci, value=float(v) if v is not None and hasattr(v, "__float__") else v)

    last_data_row = len(result) + 1
    total_row = ["合计"] + [""] * (len(visible) - 1)
    for ci, key in sum_cols:
        sm = 0
        for row in result:
            v = row.get(key)
            if v is not None and hasattr(v, "__float__"):
                sm += float(v)
        if sm:
            total_row[ci] = sm
    ws.append([])
    ws.append(total_row)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"{req.site}_{req.table_name}_sum.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/export-detail")
async def export_stats_detail(req: StatsRequest):
    """Export log detail as a separate xlsx file (chunked streaming)."""
    config_req = type("Config", (), {"db_name": f"sum_{req.site}", "table_name": req.table_name})()
    detail_cols = _build_detail_columns(req.show_channel_name, req.fields)
    if not detail_cols:
        return Response(content=b"", status_code=204)

    where, params = _detail_where(req.filters)

    wb = openpyxl.Workbook(write_only=True)
    await _write_detail_sheets(wb, config_req, detail_cols, where, params)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"{req.site}_{req.table_name}_明细.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/export-detail-async")
async def export_stats_detail_async(req: StatsRequest):
    """Start background detail export and return task_id."""
    filters_dict = req.filters
    task_id = start_stats_detail_task(
        req.site, req.table_name, filters_dict,
        req.show_channel_name, req.fields
    )
    return {"task_id": task_id}


@router.get("/export-detail-status")
async def export_stats_detail_status(task_id: str = Query(...)):
    result = get_stats_detail_status(task_id)
    if not result:
        raise HTTPException(404, detail="任务不存在")
    return result


@router.get("/export-detail-download")
async def export_stats_detail_download(task_id: str = Query(...)):
    res = get_stats_detail_download(task_id)
    if not res:
        raise HTTPException(404, detail="明细文件尚未生成或已被清理")
    path, fn = res
    return FileResponse(path, filename=fn, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
