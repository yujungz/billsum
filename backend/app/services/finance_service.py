"""Finance service - supplier reconciliation, user statistics, site monthly report."""

import os
import re
import time
import uuid
import asyncio
import calendar
import logging
import tempfile

from app import database as db
from app.config import AppConfig

log = logging.getLogger(__name__)

# ── Shared SQL fragments for cost calculation ──
# Reference: 5M cache ratio = 1.25, 1H cache ratio = 2.00
# group_ratio is applied in fee calculations but NOT in unit prices

_1H_CASE = ("CASE WHEN l.cache_creation_tokens - l.cache_creation_tokens_5m > 0 "
            "THEN l.cache_creation_tokens - l.cache_creation_tokens_5m ELSE 0 END")
_GREATEST = "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m)"


def parse_log_table_dates(table_name: str) -> tuple[str, str]:
    """Extract start/end dates from a logs table name.
    logs20260401_20260528 -> ('2026-04-01', '2026-05-28')
    logs202605 -> ('2026-05-01', '2026-05-31')
    """
    m = re.match(r'^logs(\d{8})_(\d{8})$', table_name)
    if m:
        s, e = m.group(1), m.group(2)
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}", f"{e[:4]}-{e[4:6]}-{e[6:8]}"
    m = re.match(r'^logs(\d{6})$', table_name)
    if m:
        ym = m.group(1)
        y, mo = int(ym[:4]), int(ym[4:6])
        last = calendar.monthrange(y, mo)[1]
        return f"{y:04d}-{mo:02d}-01", f"{y:04d}-{mo:02d}-{last:02d}"
    return "", ""


async def get_log_tables(site: str) -> list[str]:
    """Get list of output log tables for a site."""
    config = AppConfig.load()
    db_name = config.db_name(site)
    rows = await db.fetch_all(
        "SELECT TABLE_NAME as name FROM information_schema.TABLES "
        "WHERE TABLE_SCHEMA=%s AND TABLE_NAME LIKE 'logs%%' AND TABLE_NAME NOT LIKE '%%orig' "
        "ORDER BY TABLE_NAME",
        (db_name,),
    )
    return [r["name"] for r in rows]


async def get_usernames(site: str, table: str = "") -> list[str]:
    """Get distinct usernames. When a log table is given, query it directly for accuracy."""
    config = AppConfig.load()
    db_name = config.db_name(site)
    if table:
        rows = await db.fetch_all(
            f"SELECT DISTINCT username FROM `{table}` WHERE username IS NOT NULL AND username != '' ORDER BY username",
            db=db_name,
        )
        return [r["username"] for r in rows]
    try:
        rows = await db.fetch_all(
            "SELECT DISTINCT username FROM users WHERE username IS NOT NULL AND username != '' ORDER BY username",
            db=db_name,
        )
        if rows:
            return [r["username"] for r in rows]
    except Exception:
        pass
    return []


# ── Supplier reconciliation ──

SUPPLIER_SQL = """
SELECT
  DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d') AS `日期`,
  l.model_name AS `模型名称`,
  %s AS `供应商名称`,
  SUM(l.prompt_tokens+l.completion_tokens+l.cache_tokens+
    (CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
      THEN l.cache_creation_tokens_5m ELSE l.cache_creation_tokens END)) AS `定价单位/tokens数`,
  '美元' AS `定价币种`,
  MAX(l.model_ratio*2) AS `结算含税价input`,
  MAX(CAST(l.completion_ratio*l.model_ratio*2 AS DECIMAL(18,6))) AS `结算含税价output`,
  SUM(l.prompt_tokens) AS `输入tokens`,
  SUM(l.completion_tokens) AS `输出tokens`,
  COUNT(*) AS `调用次数`,
  SUM(l.prompt_tokens*l.model_ratio*2*l.group_ratio/1000000) AS `input费用`,
  SUM(CAST(l.completion_tokens*l.completion_ratio*l.model_ratio*2*l.group_ratio/1000000 AS DECIMAL(18,6))) AS `output费用`,
  SUM(CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
    THEN l.cache_creation_tokens_5m ELSE l.cache_creation_tokens END) AS `缓存创建`,
  SUM(l.cache_tokens) AS `缓存读取`,
  MAX(CAST((CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
    THEN l.cache_creation_ratio_5m ELSE l.cache_creation_ratio END)*l.model_ratio*2 AS DECIMAL(18,6))) AS `缓存创建单价`,
  MAX(CAST(l.cache_ratio*l.model_ratio*2 AS DECIMAL(18,6))) AS `缓存读取单价`,
  SUM((CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
    THEN l.cache_creation_tokens_5m ELSE l.cache_creation_tokens END)+l.cache_tokens) AS `cache的token`,
  SUM(CAST(((CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
      THEN l.cache_creation_tokens_5m ELSE l.cache_creation_tokens END)*
    (CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
      THEN l.cache_creation_ratio_5m ELSE l.cache_creation_ratio END)*l.model_ratio*2*l.group_ratio+
    (l.cache_tokens*cache_ratio*l.model_ratio*2*l.group_ratio))/1000000 AS DECIMAL(18,6))) AS `cache的金额`,
  SUM(CAST(((CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
      THEN l.cache_creation_tokens_5m ELSE l.cache_creation_tokens END)*
    (CASE WHEN l.cache_creation_tokens_5m > l.cache_creation_tokens
      THEN l.cache_creation_ratio_5m ELSE l.cache_creation_ratio END)*l.model_ratio*2*l.group_ratio+
    (l.cache_tokens*cache_ratio*l.model_ratio*2*l.group_ratio)+
    (l.prompt_tokens*l.model_ratio*2*l.group_ratio)+
    (l.completion_tokens*l.completion_ratio*l.model_ratio*2*l.group_ratio))
    AS DECIMAL(18,6))/1000000) AS `总费用（USD）`
FROM `{table}` l
WHERE l.windup_type < 2
  AND l.username=%s
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d'), l.model_name
ORDER BY `日期` DESC, `总费用（USD）` DESC
"""


def _date_where(date_start: str, date_end: str, alias: str = "l") -> tuple[str, list]:
    """Build optional date WHERE clause. Returns (sql_fragment, params)."""
    parts, params = [], []
    if date_start:
        parts.append(f"{alias}.created_at>=UNIX_TIMESTAMP(%s)-28800")
        params.append(f"{date_start} 00:00:00")
    if date_end:
        parts.append(f"{alias}.created_at<=UNIX_TIMESTAMP(%s)-28800")
        params.append(f"{date_end} 23:59:59")
    return (" AND " + " AND ".join(parts) if parts else ""), params


async def supplier_query(site: str, table: str, username: str,
                         date_start: str, date_end: str,
                         supplier_name: str = "") -> list[dict]:
    config = AppConfig.load()
    db_name = config.db_name(site)
    supplier_col = "%s" if supplier_name else "l.cn_supplier1"
    supplier_params = [supplier_name] if supplier_name else []
    supplier_group = "" if supplier_name else ", l.cn_supplier1"
    dw, dp = _date_where(date_start, date_end)
    sql = f"""
    SELECT
      DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d') AS `日期`,
      l.model_name AS `模型名称`,
      {supplier_col} AS `供应商名称`,
      SUM(l.prompt_tokens+l.completion_tokens+l.cache_tokens+{_GREATEST}) AS `定价单位/tokens数`,
      '美元' AS `定价币种`,
      MAX(l.model_ratio*2) AS `结算含税价input`,
      MAX(CAST(l.completion_ratio*l.model_ratio*2 AS DECIMAL(18,6))) AS `结算含税价output`,
      SUM(l.prompt_tokens) AS `输入tokens`,
      SUM(l.completion_tokens) AS `输出tokens`,
      COUNT(*) AS `调用次数`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*l.prompt_tokens/1000000 AS DECIMAL(18,6))) AS `input费用`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*l.completion_ratio*l.completion_tokens/1000000 AS DECIMAL(18,6))) AS `output费用`,
      SUM(l.cache_creation_tokens_5m) AS `创建缓存5M`,
      SUM({_1H_CASE}) AS `创建缓存1H`,
      SUM(l.cache_tokens) AS `缓存读取`,
      MAX(CAST(l.model_ratio*2*1.25 AS DECIMAL(18,6))) AS `创建缓存5M单价`,
      MAX(CAST(l.model_ratio*2*2.00 AS DECIMAL(18,6))) AS `创建缓存1H单价`,
      MAX(CAST(l.cache_ratio*l.model_ratio*2 AS DECIMAL(18,6))) AS `缓存读取单价`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*1.25*l.cache_creation_tokens_5m/1000000 AS DECIMAL(18,6))) AS `创建缓存5M费用`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*2.00*{_1H_CASE}/1000000 AS DECIMAL(18,6))) AS `创建缓存1H费用`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*l.cache_ratio*l.cache_tokens/1000000 AS DECIMAL(18,6))) AS `缓存读取费用`,
      SUM({_GREATEST}+l.cache_tokens) AS `cache的token`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
        +l.cache_ratio*l.cache_tokens
        +2.00*{_1H_CASE})/1000000 AS DECIMAL(18,6))) AS `cache的金额`,
      SUM(CAST(l.group_ratio*l.model_ratio*2*(l.prompt_tokens
        +l.completion_ratio*l.completion_tokens
        +l.cache_ratio*l.cache_tokens
        +1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000 AS DECIMAL(18,6))) AS `总费用（USD）`
    FROM `{table}` l
    WHERE l.windup_type < 2
      AND l.username=%s{dw}
    GROUP BY DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d'), l.model_name{supplier_group}
    ORDER BY `日期` DESC, `总费用（USD）` DESC
    """
    rows = await db.fetch_all(sql, supplier_params + [username] + dp, db=db_name)
    return rows


# ── User statistics ──


def _user_where(username: str) -> tuple[str, list]:
    """WHERE fragment + params for username filtering.
    Empty username → no filter (all users)."""
    if username:
        return " AND l.username=%s", [username]
    return "", []


async def user_monthly(site: str, table: str, username: str,
                       date_start: str, date_end: str,
                       show_model: bool = False) -> list[dict]:
    config = AppConfig.load()
    db_name = config.db_name(site)
    dw, dp = _date_where(date_start, date_end)
    uw, up = _user_where(username)
    extra_sel = ",\n      l.model_name AS `模型名`" if show_model else ""
    extra_grp = ", l.model_name" if show_model else ""
    sql = f"""
    SELECT
      CONCAT(DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%m'),'月份') AS `结算周期`,
      l.username AS `用户名`{extra_sel},
      SUM(l.prompt_tokens)/1000000 AS `输入token(M)`,
      SUM(l.completion_tokens)/1000000 AS `输出token(M)`,
      ROUND(SUM(l.group_ratio*l.model_ratio*2*(l.prompt_tokens
        +l.completion_ratio*l.completion_tokens
        +l.cache_ratio*l.cache_tokens
        +1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000), 6) AS `消费额度`,
      SUM(l.quota)*2/1000000 AS `平台额度`
    FROM `{table}` l
    WHERE l.windup_type < 2{uw}{dw}
    GROUP BY CONCAT(DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%m'),'月份'), l.username{extra_grp}
    ORDER BY `结算周期`, `消费额度` DESC
    """
    return await db.fetch_all(sql, up + dp, db=db_name)


async def user_daily(site: str, table: str, username: str,
                     date_start: str, date_end: str,
                     show_model: bool = False, show_token: bool = False) -> list[dict]:
    config = AppConfig.load()
    db_name = config.db_name(site)
    dw, dp = _date_where(date_start, date_end)
    uw, up = _user_where(username)
    extra_sel = ""
    extra_grp = ""
    if show_model:
        extra_sel += ",\n      l.model_name AS `模型名`"
        extra_grp += ", l.model_name"
    if show_token:
        extra_sel += ",\n      l.token_name AS `Token名称`"
        extra_grp += ", l.token_name"
    sql = f"""
    SELECT
      l.user_id AS `用户ID`,
      l.username AS `用户名`{extra_sel},
      DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d') AS `日期`,
      CONCAT(DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%m'),'月份') AS `结算周期`,
      SUM(l.prompt_tokens)/1000000 AS `输入token(M)`,
      SUM(l.completion_tokens)/1000000 AS `输出token(M)`,
      SUM(l.cache_creation_tokens_5m)/1000000 AS `创建缓存5M(M)`,
      SUM({_1H_CASE})/1000000 AS `创建缓存1H(M)`,
      SUM(l.cache_tokens)/1000000 AS `读取缓存(M)`,
      ROUND(SUM(l.group_ratio*l.model_ratio*2*(l.prompt_tokens
        +l.completion_ratio*l.completion_tokens
        +l.cache_ratio*l.cache_tokens
        +1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000), 6) AS `消费额度`,
      SUM(l.quota)*2/1000000 AS `平台额度`
    FROM `{table}` l
    WHERE l.windup_type < 2{uw}{dw}
    GROUP BY l.user_id, l.username{extra_grp},
      DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d'),
      CONCAT(DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%m'),'月份')
    ORDER BY `日期` DESC, `消费额度` DESC
    """
    return await db.fetch_all(sql, up + dp, db=db_name)


async def user_detail(site: str, table: str, username: str,
                      date_start: str, date_end: str,
                      page: int = 1, page_size: int = 20) -> dict:
    config = AppConfig.load()
    db_name = config.db_name(site)
    dw, dp = _date_where(date_start, date_end)
    uw, up = _user_where(username)

    # Count total
    count_sql = f"SELECT COUNT(*) as total FROM `{{table}}` l WHERE l.windup_type < 2{{uw}}{{dw}}"
    count_sql = count_sql.format(table=table, uw=uw, dw=dw)
    count_row = await db.fetch_one(count_sql, up + dp, db=db_name)
    total = count_row["total"] if count_row else 0

    # Query page
    offset = (page - 1) * page_size
    sql = f"""
    SELECT
      l.id AS `序号`,
      l.user_id AS `用户ID`,
      l.username AS `用户名`,
      l.token_name AS `Token名称`,
      l.model_name AS `模型名`,
      DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d %%H:%%i:%%s') AS `时间`,
      l.prompt_tokens AS `输入token`,
      ROUND(l.model_ratio*2, 6) AS `输入单价`,
      ROUND(l.group_ratio*l.model_ratio*2*l.prompt_tokens/1000000, 6) AS `输入费用`,
      l.completion_tokens AS `输出token`,
      ROUND(l.model_ratio*2*l.completion_ratio, 6) AS `输出单价`,
      ROUND(l.group_ratio*l.model_ratio*2*l.completion_ratio*l.completion_tokens/1000000, 6) AS `输出费用`,
      l.cache_tokens AS `读取缓存token`,
      ROUND(l.model_ratio*2*l.cache_ratio, 6) AS `读取缓存单价`,
      ROUND(l.group_ratio*l.model_ratio*2*l.cache_ratio*l.cache_tokens/1000000, 6) AS `读取缓存费用`,
      l.cache_creation_tokens_5m AS `创建缓存5M-token`,
      ROUND(l.model_ratio*2*1.25, 6) AS `创建缓存5M单价`,
      ROUND(l.group_ratio*l.model_ratio*2*1.25*l.cache_creation_tokens_5m/1000000, 6) AS `创建缓存5M费用`,
      {_1H_CASE} AS `创建缓存1H-token`,
      ROUND(l.model_ratio*2*2.00, 6) AS `创建缓存1H单价`,
      ROUND(l.group_ratio*l.model_ratio*2*2.00*{_1H_CASE}/1000000, 6) AS `创建缓存1H费用`,
      {_GREATEST} AS `创建缓存token`,
      ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000, 6) AS `创建缓存费用`,
      {_GREATEST}+l.cache_tokens AS `缓存总token`,
      ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
        +l.cache_ratio*l.cache_tokens
        +2.00*{_1H_CASE})/1000000, 6) AS `缓存总费用`,
      {_GREATEST}+l.cache_tokens+l.completion_tokens+l.prompt_tokens AS `总消耗token`,
      ROUND(l.group_ratio*l.model_ratio*2*(l.prompt_tokens
        +l.completion_ratio*l.completion_tokens
        +l.cache_ratio*l.cache_tokens
        +1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000, 6) AS `消费额度`,
      l.quota*2/1000000 AS `平台额度`
    FROM `{{table}}` l
    WHERE l.windup_type < 2{{uw}}{{dw}}
    ORDER BY l.created_at DESC
    LIMIT %s OFFSET %s
    """
    sql = sql.format(table=table, uw=uw, dw=dw)
    rows = await db.fetch_all(sql, up + dp + [page_size, offset], db=db_name)

    # Aggregate totals across ALL records (not just current page)
    totals_sql = f"""
    SELECT
      SUM(l.prompt_tokens) AS `输入token`,
      SUM(l.group_ratio*l.model_ratio*2*l.prompt_tokens/1000000) AS `输入费用`,
      SUM(l.completion_tokens) AS `输出token`,
      SUM(l.group_ratio*l.model_ratio*2*l.completion_ratio*l.completion_tokens/1000000) AS `输出费用`,
      SUM(l.cache_tokens) AS `读取缓存token`,
      SUM(l.group_ratio*l.model_ratio*2*l.cache_ratio*l.cache_tokens/1000000) AS `读取缓存费用`,
      SUM(l.cache_creation_tokens_5m) AS `创建缓存5M-token`,
      SUM(l.group_ratio*l.model_ratio*2*1.25*l.cache_creation_tokens_5m/1000000) AS `创建缓存5M费用`,
      SUM({_1H_CASE}) AS `创建缓存1H-token`,
      SUM(l.group_ratio*l.model_ratio*2*2.00*{_1H_CASE}/1000000) AS `创建缓存1H费用`,
      SUM({_GREATEST}) AS `创建缓存token`,
      SUM(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000) AS `创建缓存费用`,
      SUM({_GREATEST}+l.cache_tokens) AS `缓存总token`,
      SUM(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
        +l.cache_ratio*l.cache_tokens
        +2.00*{_1H_CASE})/1000000) AS `缓存总费用`,
      SUM({_GREATEST}+l.cache_tokens+l.completion_tokens+l.prompt_tokens) AS `总消耗token`,
      SUM(l.group_ratio*l.model_ratio*2*(l.prompt_tokens
        +l.completion_ratio*l.completion_tokens
        +l.cache_ratio*l.cache_tokens
        +1.25*l.cache_creation_tokens_5m
        +2.00*{_1H_CASE})/1000000) AS `消费额度`,
      SUM(l.quota*2/1000000) AS `平台额度`
    FROM `{{table}}` l
    WHERE l.windup_type < 2{{uw}}{{dw}}
    """
    totals_sql = totals_sql.format(table=table, uw=uw, dw=dw)
    totals_row = await db.fetch_one(totals_sql, up + dp, db=db_name)
    totals = {}
    if totals_row:
        for k, v in totals_row.items():
            totals[k] = float(v) if v is not None and hasattr(v, "__float__") else v

    return {"total": total, "page": page, "page_size": page_size, "data": rows, "totals": totals}


# ── Async export tasks ──

_export_tasks: dict[str, dict] = {}

_CHUNK_SIZE = 50000       # rows per DB query
_MAX_ROWS_PER_SHEET = 1000000


def start_export_task(site: str, table: str, username: str,
                      date_start: str, date_end: str,
                      with_platform: bool, with_detail: bool = True,
                      show_model: bool = False, show_token: bool = False,
                      with_total_cost: bool = True) -> str:
    """Start a background export task, return task_id immediately."""
    task_id = uuid.uuid4().hex[:8]
    _export_tasks[task_id] = {
        "task_id": task_id,
        "status": "running",
        "progress": "",
        "site": site,
        "table": table,
        "username": username,
        "date_start": date_start,
        "date_end": date_end,
        "start_time": time.time(),
        "end_time": None,
        "file_path": None,
        "error": None,
    }

    async def _run():
        try:
            await _run_export(task_id, site, table, username, date_start, date_end,
                              with_platform, with_detail, show_model, show_token,
                              with_total_cost)
        except Exception as e:
            log.exception(f"[export-{task_id}] Failed: {e}")
            _export_tasks[task_id]["status"] = "failed"
            _export_tasks[task_id]["error"] = str(e)
            _export_tasks[task_id]["end_time"] = time.time()

    asyncio.create_task(_run())
    return task_id


def get_export_status(task_id: str) -> dict | None:
    t = _export_tasks.get(task_id)
    if not t:
        return None
    elapsed = ((t["end_time"] or time.time()) - t["start_time"])
    return {**t, "elapsed": round(elapsed, 1)}


def cleanup_export_task(task_id: str):
    """Remove task entry and its temp file."""
    t = _export_tasks.pop(task_id, None)
    if t and t.get("file_path"):
        try:
            os.remove(t["file_path"])
        except OSError:
            pass


async def _run_export(task_id: str, site: str, table: str, username: str,
                      date_start: str, date_end: str, with_platform: bool,
                      with_detail: bool = True,
                      show_model: bool = False, show_token: bool = False,
                      with_total_cost: bool = True):
    """Generate xlsx file with streaming writes and multi-sheet splitting."""
    import openpyxl
    from openpyxl.styles import Font, Alignment

    task = _export_tasks[task_id]
    config = AppConfig.load()
    db_name = config.db_name(site)
    dw, dp = _date_where(date_start, date_end)

    # 1. Fetch monthly & daily (small datasets)
    task["progress"] = "查询汇总数据..."
    monthly = await user_monthly(site, table, username, date_start, date_end, show_model)
    daily = await user_daily(site, table, username, date_start, date_end, show_model, show_token)

    # 2. Count detail rows (skipped entirely when detail not requested)
    if not with_detail:
        detail_total = 0
    else:
        uw, up = _user_where(username)
        count_sql = f"SELECT COUNT(*) AS total FROM `{table}` l WHERE l.windup_type < 2{uw}{dw}"
        count_row = await db.fetch_one(count_sql, up + dp, db=db_name)
        detail_total = count_row["total"] if count_row else 0

    # 3. Build Excel with write-only workbook
    _PLATFORM_KEYS = {"平台额度"}
    if not with_platform:
        monthly = _strip_platform(monthly, _PLATFORM_KEYS)
        daily = _strip_platform(daily, _PLATFORM_KEYS)
    if not with_total_cost:
        monthly = _strip_platform(monthly, {"消费额度"})
        daily = _strip_platform(daily, {"消费额度"})

    total_summary = (["消费额度"] if with_total_cost else []) + (["平台额度"] if with_platform else [])
    total_detail_fields = (["消费额度"] if with_total_cost else []) + (["平台额度"] if with_platform else [])

    wb = openpyxl.Workbook(write_only=True)

    # Monthly sheet (small)
    if monthly:
        ws_month = wb.create_sheet("月汇总")
        _write_sheet_streaming(ws_month, monthly, total_summary)

    # Daily sheet (small)
    if daily:
        ws_daily = wb.create_sheet("日统计")
        _write_sheet_streaming(ws_daily, daily, total_summary)

    # Detail sheets — chunked reads, multi-sheet splitting
    if detail_total > 0:
        task["progress"] = f"导出明细 0/{detail_total}"

        detail_sql = f"""
        SELECT
          l.id AS `序号`,
          l.user_id AS `用户ID`,
          l.username AS `用户名`,
          l.token_name AS `Token名称`,
          l.model_name AS `模型名`,
          DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d %%H:%%i:%%s') AS `时间`,
          l.prompt_tokens AS `输入token`,
          ROUND(l.model_ratio*2, 6) AS `输入单价`,
          ROUND(l.group_ratio*l.model_ratio*2*l.prompt_tokens/1000000, 6) AS `输入费用`,
          l.completion_tokens AS `输出token`,
          ROUND(l.model_ratio*2*l.completion_ratio, 6) AS `输出单价`,
          ROUND(l.group_ratio*l.model_ratio*2*l.completion_ratio*l.completion_tokens/1000000, 6) AS `输出费用`,
          l.cache_tokens AS `读取缓存token`,
          ROUND(l.model_ratio*2*l.cache_ratio, 6) AS `读取缓存单价`,
          ROUND(l.group_ratio*l.model_ratio*2*l.cache_ratio*l.cache_tokens/1000000, 6) AS `读取缓存费用`,
          l.cache_creation_tokens_5m AS `创建缓存5M-token`,
          ROUND(l.model_ratio*2*1.25, 6) AS `创建缓存5M单价`,
          ROUND(l.group_ratio*l.model_ratio*2*1.25*l.cache_creation_tokens_5m/1000000, 6) AS `创建缓存5M费用`,
          {_1H_CASE} AS `创建缓存1H-token`,
          ROUND(l.model_ratio*2*2.00, 6) AS `创建缓存1H单价`,
          ROUND(l.group_ratio*l.model_ratio*2*2.00*{_1H_CASE}/1000000, 6) AS `创建缓存1H费用`,
          {_GREATEST} AS `创建缓存token`,
          ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
            +2.00*{_1H_CASE})/1000000, 6) AS `创建缓存费用`,
          {_GREATEST}+l.cache_tokens AS `缓存总token`,
          ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m
            +l.cache_ratio*l.cache_tokens
            +2.00*{_1H_CASE})/1000000, 6) AS `缓存总费用`,
          {_GREATEST}+l.cache_tokens+l.completion_tokens+l.prompt_tokens AS `总消耗token`,
          ROUND(l.group_ratio*l.model_ratio*2*(l.prompt_tokens
            +l.completion_ratio*l.completion_tokens
            +l.cache_ratio*l.cache_tokens
            +1.25*l.cache_creation_tokens_5m
            +2.00*{_1H_CASE})/1000000, 6) AS `消费额度`,
          l.quota*2/1000000 AS `平台额度`
        FROM `{table}` l
        WHERE l.windup_type < 2{uw}{dw}
        ORDER BY l.created_at DESC
        LIMIT %s OFFSET %s
        """

        # Strip platform columns from detail if needed
        strip_keys = {"平台额度"} if not with_platform else set()
        headers_written = False
        headers = None
        ws_detail = None
        sheet_idx = 0
        rows_in_sheet = 0
        exported = 0

        offset = 0
        while offset < detail_total:
            chunk = await db.fetch_all(
                detail_sql, up + dp + [_CHUNK_SIZE, offset], db=db_name
            )
            if not chunk:
                break

            for row in chunk:
                # Start new sheet if needed
                if ws_detail is None or rows_in_sheet >= _MAX_ROWS_PER_SHEET:
                    sheet_idx += 1
                    sheet_name = "用户明细" if sheet_idx == 1 else f"用户明细_{sheet_idx}"
                    ws_detail = wb.create_sheet(sheet_name)
                    headers_written = False
                    rows_in_sheet = 0

                if not headers_written:
                    headers = [k for k in row.keys() if k not in strip_keys]
                    # Header row with bold style
                    header_cells = [openpyxl.cell.Cell(ws_detail, column=i+1, value=h) for i, h in enumerate(headers)]
                    for c in header_cells:
                        c.font = Font(bold=True)
                        c.alignment = Alignment(horizontal="center")
                    ws_detail.append(header_cells)
                    headers_written = True

                vals = [row[k] for k in headers]
                ws_detail.append(vals)
                rows_in_sheet += 1

            exported += len(chunk)
            offset += _CHUNK_SIZE
            task["progress"] = f"导出明细 {exported}/{detail_total}"

        # Append total row to last detail sheet
        if ws_detail and headers:
            # Collect index of total fields
            total_row_vals = ["合计"] + ["" for _ in range(len(headers) - 1)]
            for field in total_detail_fields:
                if field in headers:
                    ci = headers.index(field) + 1
                    cl = openpyxl.utils.get_column_letter(ci)
                    total_row_vals[ci - 1] = f"=SUM({cl}2:{cl}{rows_in_sheet + 1})"
            ws_detail.append([])
            ws_detail.append(total_row_vals)

    # 4. Save to temp file
    tmp_dir = os.path.join(tempfile.gettempdir(), "billsum_export")
    os.makedirs(tmp_dir, exist_ok=True)
    file_path = os.path.join(tmp_dir, f"{task_id}.xlsx")
    log.info(f"[export-{task_id}] Saving xlsx to {file_path}, monthly={len(monthly)}, daily={len(daily)}, detail={detail_total}")
    wb.save(file_path)
    file_size = os.path.getsize(file_path)
    log.info(f"[export-{task_id}] Saved, file_size={file_size}")

    task["status"] = "done"
    task["file_path"] = file_path
    task["end_time"] = time.time()
    task["progress"] = f"完成 ({detail_total} 条明细)"


def _write_sheet_streaming(ws, rows: list[dict], total_fields: list[str]):
    """Write rows to a write-only sheet with header bold + total row."""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    if not rows:
        return
    headers = list(rows[0].keys())
    header_cells = [openpyxl.cell.Cell(ws, column=i+1, value=h) for i, h in enumerate(headers)]
    for c in header_cells:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    ws.append(header_cells)
    for r in rows:
        ws.append(list(r.values()))

    last_data_row = len(rows) + 1
    total_row = ["合计"] + ["" for _ in range(len(headers) - 1)]
    for field in total_fields:
        if field in headers:
            ci = headers.index(field) + 1
            cl = openpyxl.utils.get_column_letter(ci)
            total_row[ci - 1] = f"=SUM({cl}2:{cl}{last_data_row})"
    ws.append([])
    ws.append(total_row)


def _strip_platform(rows, strip_keys):
    """Remove specified columns from each row dict."""
    if not rows:
        return rows
    return [{k: v for k, v in r.items() if k not in strip_keys} for r in rows]


# ── Site monthly report ──

# Channel discount: use cn_discount_orig when windup_type=1 and orig > 0, else cn_discount
CN_DISCOUNT_EXPR = """CASE
    WHEN l.windup_type = 1 AND l.cn_discount_orig IS NOT NULL AND l.cn_discount_orig > 0
    THEN l.cn_discount_orig
    ELSE l.cn_discount
END"""

_1H = "CASE WHEN l.cache_creation_tokens - l.cache_creation_tokens_5m > 0 THEN l.cache_creation_tokens - l.cache_creation_tokens_5m ELSE 0 END"
_GR = "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m)"
_COST_EXPR = "l.group_ratio*l.model_ratio*2*(l.prompt_tokens+l.completion_ratio*l.completion_tokens+l.cache_ratio*l.cache_tokens+1.25*l.cache_creation_tokens_5m+2.00*"+_1H+")/1000000"
_US_D = "COALESCE(l.us_discount,0)"

PURCHASE_DETAIL_SQL = f"""
SELECT
  l.channel_id AS `渠道ID`,
  l.channel_name AS `渠道名称`,
  COUNT(*) AS `记录数`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`,
  ROUND(SUM({_US_D}*{_COST_EXPR}), 6) AS `收入`,
  ROUND(SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `成本`,
  ROUND(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `毛利`,
  ROUND({{purchase_rate}}*(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR})), 6) AS `采购提成`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.cn_buyer1 = %s
  AND l.cn_supplier1 = %s
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.channel_id, l.channel_name
"""

PURCHASE_SUMMARY_SQL = f"""
SELECT
  l.cn_supplier1 AS `供应商`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`,
  ROUND(SUM({_US_D}*{_COST_EXPR}), 6) AS `收入`,
  ROUND(SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `成本`,
  ROUND(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `毛利`,
  ROUND({{purchase_rate}}*(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR})), 6) AS `采购提成`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.cn_buyer1 = %s
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.cn_supplier1
"""

SALES_DETAIL_SQL = f"""
SELECT
  l.user_id AS `用户ID`,
  l.username AS `用户名`,
  l.token_name AS `Token名称`,
  l.channel_id AS `渠道ID`,
  l.`group` AS `用户组`,
  DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d') AS `日期`,
  SUM(l.prompt_tokens)/1000000 AS `输入token(M)`,
  SUM(l.completion_tokens)/1000000 AS `输出token(M)`,
  SUM(l.cache_tokens)/1000000 AS `读取缓存token(M)`,
  SUM({_GR})/1000000 AS `创建缓存token(M)`,
  SUM(l.cache_creation_tokens_5m)/1000000 AS `创建缓存5M-token(M)`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`,
  MAX(l.us_discount) AS `用户折扣`,
  MAX({{cn_discount}}) AS `渠道折扣`,
  ROUND(SUM({_US_D}*{_COST_EXPR}), 6) AS `收入`,
  ROUND(SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `成本`,
  ROUND(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `毛利`,
  ROUND({{sales_rate}}*(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR})), 6) AS `提成`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.us_salesperson = %s
  AND l.username = %s
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.user_id, l.username, l.token_name, l.channel_id, l.`group`,
  DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d')
ORDER BY `日期` DESC
"""

SALES_SUMMARY_SQL = f"""
SELECT
  l.user_id AS `用户ID`,
  l.username AS `用户名`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`,
  ROUND(SUM({_US_D}*{_COST_EXPR}), 6) AS `收入`,
  ROUND(SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `成本`,
  ROUND(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `毛利`,
  ROUND({{sales_rate}}*(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR})), 6) AS `提成`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.us_salesperson = %s
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.user_id, l.username
"""

CUSTOMER_DETAIL_SQL = f"""
SELECT
  l.user_id AS `用户ID`,
  l.username AS `用户名`,
  l.token_name AS `Token名称`,
  l.`group` AS `用户组`,
  DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d') AS `日期`,
  SUM(l.prompt_tokens)/1000000 AS `输入token(M)`,
  SUM(l.completion_tokens)/1000000 AS `输出token(M)`,
  SUM(l.cache_tokens)/1000000 AS `读取缓存token(M)`,
  SUM({_GR})/1000000 AS `创建缓存token(M)`,
  SUM(l.cache_creation_tokens_5m)/1000000 AS `创建缓存5M-token(M)`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.us_salesperson = %s
  AND l.username = %s
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.user_id, l.username, l.token_name, l.`group`,
  DATE_FORMAT(FROM_UNIXTIME(l.created_at+28800), '%%Y-%%m-%%d')
ORDER BY `日期` DESC
"""

PURCHASE_OVERVIEW_SQL = f"""
SELECT
  l.cn_buyer1 AS `采购员`,
  COUNT(DISTINCT l.cn_supplier1) AS `供应商数`,
  COUNT(*) AS `记录数`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`,
  ROUND(SUM({_US_D}*{_COST_EXPR}), 6) AS `收入`,
  ROUND(SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `成本`,
  ROUND(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `毛利`,
  ROUND({{purchase_rate}}*(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR})), 6) AS `采购提成`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.cn_buyer1 IS NOT NULL AND l.cn_buyer1 != ''
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.cn_buyer1
"""

SALES_OVERVIEW_SQL = f"""
SELECT
  l.us_salesperson AS `销售员`,
  COUNT(DISTINCT l.username) AS `用户数`,
  COUNT(*) AS `记录数`,
  ROUND(SUM({_COST_EXPR}), 6) AS `消费额度`,
  ROUND(SUM({_US_D}*{_COST_EXPR}), 6) AS `收入`,
  ROUND(SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `成本`,
  ROUND(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR}), 6) AS `毛利`,
  ROUND({{sales_rate}}*(SUM({_US_D}*{_COST_EXPR}) - SUM({{cn_discount}}*{_COST_EXPR})), 6) AS `提成`
FROM `{{table}}` l
WHERE l.windup_type < 2
  AND l.us_salesperson IS NOT NULL AND l.us_salesperson != ''
  AND l.created_at>=UNIX_TIMESTAMP(%s)-28800
  AND l.created_at<=UNIX_TIMESTAMP(%s)-28800
GROUP BY l.us_salesperson
"""


def _get_commission_rates() -> tuple[float, float]:
    config = AppConfig.load()
    return config.business.purchase_commission_rate, config.business.sales_commission_rate


async def check_ex_tables(site: str) -> bool:
    config = AppConfig.load()
    db_name = config.db_name(site)
    try:
        ch = await db.fetch_one("SELECT COUNT(*) as cnt FROM ex_channels", db=db_name)
        us = await db.fetch_one("SELECT COUNT(*) as cnt FROM ex_users", db=db_name)
        return (ch and ch["cnt"] > 0) and (us and us["cnt"] > 0)
    except Exception:
        return False


async def _get_buyers(db_name: str, table: str, ds: str, de: str) -> list[str]:
    rows = await db.fetch_all(
        f"SELECT DISTINCT cn_buyer1 FROM `{table}` l "
        "WHERE l.windup_type < 2 AND cn_buyer1 IS NOT NULL AND cn_buyer1 != '' "
        "AND l.created_at>=UNIX_TIMESTAMP(%s)-28800 AND l.created_at<=UNIX_TIMESTAMP(%s)-28800 "
        "ORDER BY cn_buyer1",
        (ds, de), db=db_name,
    )
    return [r["cn_buyer1"] for r in rows]


async def _get_suppliers(db_name: str, table: str, buyer: str, ds: str, de: str) -> list[str]:
    rows = await db.fetch_all(
        f"SELECT DISTINCT cn_supplier1 FROM `{table}` l "
        "WHERE l.windup_type < 2 AND l.cn_buyer1=%s "
        "AND cn_supplier1 IS NOT NULL AND cn_supplier1 != '' "
        "AND l.created_at>=UNIX_TIMESTAMP(%s)-28800 AND l.created_at<=UNIX_TIMESTAMP(%s)-28800 "
        "ORDER BY cn_supplier1",
        (buyer, ds, de), db=db_name,
    )
    return [r["cn_supplier1"] for r in rows]


async def _get_salespeople(db_name: str, table: str, ds: str, de: str) -> list[str]:
    rows = await db.fetch_all(
        f"SELECT DISTINCT us_salesperson FROM `{table}` l "
        "WHERE l.windup_type < 2 AND us_salesperson IS NOT NULL AND us_salesperson != '' "
        "AND l.created_at>=UNIX_TIMESTAMP(%s)-28800 AND l.created_at<=UNIX_TIMESTAMP(%s)-28800 "
        "ORDER BY us_salesperson",
        (ds, de), db=db_name,
    )
    return [r["us_salesperson"] for r in rows]


async def _get_users_for_sp(db_name: str, table: str, sp: str, ds: str, de: str) -> list[dict]:
    return await db.fetch_all(
        f"SELECT DISTINCT l.user_id, l.username FROM `{table}` l "
        "WHERE l.windup_type < 2 AND l.us_salesperson=%s "
        "AND l.username IS NOT NULL AND l.username != '' "
        "AND l.created_at>=UNIX_TIMESTAMP(%s)-28800 AND l.created_at<=UNIX_TIMESTAMP(%s)-28800 "
        "ORDER BY l.user_id",
        (sp, ds, de), db=db_name,
    )


def _make_workbook(headers: list[str], rows: list[dict],
                   total_fields: list[str] | None = None):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = center

    for r in rows:
        ws.append(list(r.values()))

    if total_fields:
        total_row = ["" for _ in headers]
        total_row[0] = "合计"
        for field in total_fields:
            if field in headers:
                ci = headers.index(field) + 1
                cl = get_column_letter(ci)
                total_row[ci - 1] = f"=SUM({cl}2:{cl}{len(rows)+1})"
        ws.append(total_row)

    for i, h in enumerate(headers, 1):
        max_len = len(str(h)) * 2 + 2
        for r in rows:
            v = list(r.values())[i - 1]
            if v is not None:
                max_len = max(max_len, len(str(v)) + 2)
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 30)

    return wb


def _write_excel(filepath: str, headers: list[str], rows: list[dict],
                 total_fields: list[str] | None = None):
    _make_workbook(headers, rows, total_fields).save(filepath)


def _excel_bytes(headers: list[str], rows: list[dict],
                 total_fields: list[str] | None = None) -> bytes:
    from io import BytesIO
    buf = BytesIO()
    _make_workbook(headers, rows, total_fields).save(buf)
    return buf.getvalue()


async def site_report_preview(site: str, table: str, date_start: str, date_end: str) -> dict:
    config = AppConfig.load()
    db_name = config.db_name(site)
    ds = f"{date_start} 00:00:00"
    de = f"{date_end} 23:59:59"
    purchase_rate, sales_rate = _get_commission_rates()

    purchase, sales = [], []
    try:
        purchase = await db.fetch_all(
            PURCHASE_OVERVIEW_SQL.format(table=table, purchase_rate=purchase_rate, cn_discount=CN_DISCOUNT_EXPR),
            (ds, de), db=db_name,
        )
    except Exception:
        pass
    try:
        sales = await db.fetch_all(
            SALES_OVERVIEW_SQL.format(table=table, sales_rate=sales_rate, cn_discount=CN_DISCOUNT_EXPR),
            (ds, de), db=db_name,
        )
    except Exception:
        pass
    return {"purchase": purchase, "sales": sales}


async def generate_all_reports(site: str, table: str, date_start: str, date_end: str,
                               output_root: str) -> dict:
    config = AppConfig.load()
    db_name = config.db_name(site)
    ds = f"{date_start} 00:00:00"
    de = f"{date_end} 23:59:59"
    ym = date_start.replace("-", "")[:6]
    purchase_rate, sales_rate = _get_commission_rates()

    report_dir = os.path.join(output_root, f"{site}_report{ym}")
    generated = []

    # ── Purchase commission ──
    purchase_dir = os.path.join(report_dir, "采购提成")
    buyers = await _get_buyers(db_name, table, ds, de)

    for buyer in buyers:
        suppliers = await _get_suppliers(db_name, table, buyer, ds, de)
        summary_rows = await db.fetch_all(
            PURCHASE_SUMMARY_SQL.format(table=table, purchase_rate=purchase_rate, cn_discount=CN_DISCOUNT_EXPR),
            (buyer, ds, de), db=db_name,
        )
        if not suppliers and not summary_rows:
            continue

        buyer_dir = os.path.join(purchase_dir, buyer)
        os.makedirs(buyer_dir, exist_ok=True)

        for supplier in suppliers:
            rows = await db.fetch_all(
                PURCHASE_DETAIL_SQL.format(table=table, purchase_rate=purchase_rate, cn_discount=CN_DISCOUNT_EXPR),
                (buyer, supplier, ds, de), db=db_name,
            )
            if rows:
                headers = list(rows[0].keys())
                _write_excel(
                    os.path.join(buyer_dir, f"{supplier}_{ym}.xlsx"),
                    headers, rows,
                    total_fields=["记录数", "消费额度", "收入", "成本", "毛利", "采购提成"],
                )
                generated.append(f"采购提成/{buyer}/{supplier}_{ym}.xlsx")

        if summary_rows:
            headers = list(summary_rows[0].keys())
            _write_excel(
                os.path.join(buyer_dir, f"{buyer}_汇总_{ym}.xlsx"),
                headers, summary_rows,
                total_fields=["消费额度", "收入", "成本", "毛利", "采购提成"],
            )
            generated.append(f"采购提成/{buyer}/{buyer}_汇总_{ym}.xlsx")

    # ── Sales commission + Customer reports ──
    sales_dir = os.path.join(report_dir, "销售提成")
    customer_dir = os.path.join(report_dir, "客户报表")
    salespeople = await _get_salespeople(db_name, table, ds, de)

    for sp in salespeople:
        users = await _get_users_for_sp(db_name, table, sp, ds, de)
        summary_rows = await db.fetch_all(
            SALES_SUMMARY_SQL.format(table=table, sales_rate=sales_rate, cn_discount=CN_DISCOUNT_EXPR),
            (sp, ds, de), db=db_name,
        )
        if not users and not summary_rows:
            continue

        sp_sales_dir = os.path.join(sales_dir, sp)
        sp_cust_dir = os.path.join(customer_dir, sp)
        os.makedirs(sp_sales_dir, exist_ok=True)
        os.makedirs(sp_cust_dir, exist_ok=True)

        for u in users:
            uid, uname = u["user_id"], u["username"]
            fname = f"{uid}_{uname}_{ym}.xlsx"

            # Sales detail
            rows = await db.fetch_all(
                SALES_DETAIL_SQL.format(table=table, sales_rate=sales_rate, cn_discount=CN_DISCOUNT_EXPR),
                (sp, uname, ds, de), db=db_name,
            )
            if rows:
                _write_excel(
                    os.path.join(sp_sales_dir, fname),
                    list(rows[0].keys()), rows,
                    total_fields=["消费额度", "收入", "成本", "毛利", "提成"],
                )
                generated.append(f"销售提成/{sp}/{fname}")

            # Customer report
            cust_rows = await db.fetch_all(
                CUSTOMER_DETAIL_SQL.format(table=table),
                (sp, uname, ds, de), db=db_name,
            )
            if cust_rows:
                _write_excel(
                    os.path.join(sp_cust_dir, fname),
                    list(cust_rows[0].keys()), cust_rows,
                    total_fields=["消费额度"],
                )
                generated.append(f"客户报表/{sp}/{fname}")

        if summary_rows:
            _write_excel(
                os.path.join(sp_sales_dir, f"{sp}_汇总_{ym}.xlsx"),
                list(summary_rows[0].keys()), summary_rows,
                total_fields=["消费额度", "收入", "成本", "毛利", "提成"],
            )
            generated.append(f"销售提成/{sp}/{sp}_汇总_{ym}.xlsx")

    return {"report_dir": report_dir, "files": generated, "total_files": len(generated)}


async def generate_reports_zip(site: str, table: str, date_start: str, date_end: str) -> bytes:
    """Generate all reports and return as a ZIP file."""
    import zipfile
    from io import BytesIO

    config = AppConfig.load()
    db_name = config.db_name(site)
    ds = f"{date_start} 00:00:00"
    de = f"{date_end} 23:59:59"
    ym = date_start.replace("-", "")[:6]
    purchase_rate, sales_rate = _get_commission_rates()

    report_prefix = f"{site}_report{ym}"
    buf = BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # ── Purchase commission ──
        try:
            buyers = await _get_buyers(db_name, table, ds, de)
        except Exception:
            buyers = []
        for buyer in buyers:
            suppliers = await _get_suppliers(db_name, table, buyer, ds, de)
            summary_rows = await db.fetch_all(
                PURCHASE_SUMMARY_SQL.format(table=table, purchase_rate=purchase_rate, cn_discount=CN_DISCOUNT_EXPR),
                (buyer, ds, de), db=db_name,
            )
            if not suppliers and not summary_rows:
                continue
            buyer_prefix = f"{report_prefix}/采购提成/{buyer}"
            for supplier in suppliers:
                rows = await db.fetch_all(
                    PURCHASE_DETAIL_SQL.format(table=table, purchase_rate=purchase_rate, cn_discount=CN_DISCOUNT_EXPR),
                    (buyer, supplier, ds, de), db=db_name,
                )
                if rows:
                    headers = list(rows[0].keys())
                    zf.writestr(
                        f"{buyer_prefix}/{supplier}_{ym}.xlsx",
                        _excel_bytes(headers, rows, total_fields=["记录数", "消费额度", "收入", "成本", "毛利", "采购提成"]),
                    )
            if summary_rows:
                headers = list(summary_rows[0].keys())
                zf.writestr(
                    f"{buyer_prefix}/{buyer}_汇总_{ym}.xlsx",
                    _excel_bytes(headers, summary_rows, total_fields=["消费额度", "收入", "成本", "毛利", "采购提成"]),
                )

        # ── Sales commission + Customer reports ──
        try:
            salespeople = await _get_salespeople(db_name, table, ds, de)
        except Exception:
            salespeople = []
        for sp in salespeople:
            users = await _get_users_for_sp(db_name, table, sp, ds, de)
            summary_rows = await db.fetch_all(
                SALES_SUMMARY_SQL.format(table=table, sales_rate=sales_rate, cn_discount=CN_DISCOUNT_EXPR),
                (sp, ds, de), db=db_name,
            )
            if not users and not summary_rows:
                continue
            sp_sales_prefix = f"{report_prefix}/销售提成/{sp}"
            sp_cust_prefix = f"{report_prefix}/客户报表/{sp}"
            for u in users:
                uid, uname = u["user_id"], u["username"]
                fname = f"{uid}_{uname}_{ym}.xlsx"
                rows = await db.fetch_all(
                    SALES_DETAIL_SQL.format(table=table, sales_rate=sales_rate, cn_discount=CN_DISCOUNT_EXPR),
                    (sp, uname, ds, de), db=db_name,
                )
                if rows:
                    zf.writestr(
                        f"{sp_sales_prefix}/{fname}",
                        _excel_bytes(list(rows[0].keys()), rows, total_fields=["消费额度", "收入", "成本", "毛利", "提成"]),
                    )
                cust_rows = await db.fetch_all(
                    CUSTOMER_DETAIL_SQL.format(table=table),
                    (sp, uname, ds, de), db=db_name,
                )
                if cust_rows:
                    zf.writestr(
                        f"{sp_cust_prefix}/{fname}",
                        _excel_bytes(list(cust_rows[0].keys()), cust_rows, total_fields=["消费额度"]),
                    )
            if summary_rows:
                zf.writestr(
                    f"{sp_sales_prefix}/{sp}_汇总_{ym}.xlsx",
                    _excel_bytes(list(summary_rows[0].keys()), summary_rows, total_fields=["消费额度", "收入", "成本", "毛利", "提成"]),
                )

    return buf.getvalue()
