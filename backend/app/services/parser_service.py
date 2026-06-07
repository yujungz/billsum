"""Parser service - parse channels/tokens/users tables into ex_ tables and export Excel."""

import io
import re
import logging

import openpyxl
from openpyxl.styles import Font

from app.config import AppConfig
from app import database as db

log = logging.getLogger(__name__)

USD_RATIO = 6.91
RATE_BASE = 0.144356 / 3

MODEL_NAMES = [
    'claude', 'gpt', 'gemini', 'gpt-image', 'kimi', 'kfc', 'sora', 'deepseek', 'qwen',
    'openrouter', 'veo3', 'aws', 'vertex', 'azure', 'anthropic', 'glm', 'glm-5',
]

COMMON_WORDS = ['api', 'top', 'com', 'cn', 'net', 'org', 'io', 'app', 'www', 'http', 'https']

PARSEABLE_TABLES = {'channels', 'tokens', 'users'}


# ─── Channel parsing ────────────────────────────────────────────────

def _ch_parse_discount(text):
    if not text:
        return None
    text = text.strip()

    m = re.search(r'(\d+\.?\d*)\s*倍', text)
    if m:
        return round(float(m.group(1)), 2)

    m = re.search(r'(\d+\.?\d*)\s*%', text)
    if m:
        return round(float(m.group(1)) / 100, 2)

    if re.search(r'元\s*/\s*(张|次)', text):
        return 1.0

    m = re.search(r'(\d+\.?\d*)\s*折', text)
    if m:
        return round(float(m.group(1)) / 10, 2)

    m = re.search(r'(\d+\.?\d*)\s*￥\s*(一|/)?\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*￥', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*元\s*(一|/)?\s*(刀|百万)', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'^(\d+\.?\d*)\s*(一|/)?\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*(一|/)\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'^一\s*元\s*一', text)
    if m:
        return round(1.0 / USD_RATIO, 2)

    m = re.search(r'^(\d+\.?\d*)$', text)
    if m:
        val = float(m.group(1))
        if val < 10:
            return round(val / USD_RATIO, 2)
        elif val < 100:
            return round(val / 100, 2)

    m = re.search(r'(\d+\.?\d*)\s*元\s*(\d+\.?\d*)\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    return None


def _ch_is_chinese_name(text):
    if not text or len(text) < 2 or len(text) > 4:
        return False
    if not all('一' <= ch <= '鿿' for ch in text):
        return False
    discount_kw = ['￥', '折', '刀', '元', '百万', '一刀', '一折', '/刀', '/元', '一']
    if any(kw in text for kw in discount_kw):
        return False
    if any(s in text for s in ['波转', '顺']):
        return False
    if any(w in text for w in ['备注', '额外', '信息', '说明', '测试', '临时']):
        return False
    return True


def _ch_is_purchaser_candidate(text):
    if not text or len(text) > 6:
        return False
    if re.match(r'^\d', text):
        return False
    discount_kw = ['￥', '折', '刀', '元', '百万', '一刀', '一折', '/刀', '/元', '一', '%', '倍']
    if any(kw in text for kw in discount_kw):
        return False
    if re.match(r'^\d+\.?\d*$', text):
        return False
    if re.match(r'https?://', text):
        return False
    if re.match(r'^[\w.-]+\.[\w]+', text):
        return False
    site_names = ['pinova', 'csp', 'wzg', 'qn', 'ai', '波转', '顺']
    if any(s in text.lower() for s in site_names):
        return False
    supplier_names = [
        'api', 'xgapi', '4sapi', 'chatfire', 'dataeyes', 'ikuncode', 'lconai',
        'poloapi', 'poloai', 'polo', 'zeta', 'gptnb', 'wzg', 'one-ai', 'b64', 'chengfeng',
        'sunzong', 'tokshub', 'aigc', 'vip', 'aws', 'azure', 'gemini', 'claude',
        'gpt', 'deepseek', 'qwen', 'kimi', 'sora', 'veo', 'veo3', 'kfc', 'netnic',
        'heny', 'grasai', 'otuapi', 'x-see', 'submodel', 'glm', 'glm-5',
    ]
    if any(s in text.lower() for s in supplier_names):
        return False
    if any(text.lower().startswith(m) or m in text.lower() for m in MODEL_NAMES):
        return False
    if re.match(r'^\d+\.\d+\.\d+\.\d+', text):
        return False
    if any(w in text for w in ['备注', '额外', '信息', '说明', '测试', '临时']):
        return False
    return True


def _ch_clean_supplier_name(name):
    if not name:
        return 'XX'
    if re.match(r'^[\d.]+$', name) or re.match(r'^\d+\.\d+\.\d+\.\d+', name):
        return 'XX'
    discount_kw = ['￥', '折', '刀', '元', '百万', '一刀', '一折', '/刀', '/元', '%', '倍']
    if any(kw in name for kw in discount_kw):
        return 'XX'
    name = re.sub(r'^https?://', '', name)
    name = re.sub(r':\d+$', '', name)
    name = name.split('/')[0]
    name = re.sub(r'[\/!@#$%^&]', '', name)
    if '.' not in name:
        return name if name and len(name) > 1 else 'XX'
    parts = name.split('.')
    valid_parts = [p for p in parts if p.lower() not in COMMON_WORDS and len(p) > 2]
    if valid_parts:
        if len(valid_parts) >= 2:
            idx = len(parts) - 2
            sl = parts[idx] if idx >= 0 else parts[0]
            if sl.lower() not in COMMON_WORDS and len(sl) > 2:
                return sl
        return valid_parts[0]
    if len(parts) >= 2:
        sl = parts[-2]
        if sl.lower() not in COMMON_WORDS and len(sl) > 1:
            return sl
    for p in parts:
        if p.lower() not in COMMON_WORDS and len(p) > 1:
            return p
    return 'XX'


def _ch_parse_discount_and_purchaser(text):
    if not text:
        return None, None
    text = text.strip()
    # Try patterns that combine discount + possible trailing purchaser name
    # Each pattern: (regex, use_usd_ratio)
    patterns = [
        (r'(\d+\.?\d*)\s*￥\s*(一|/)?\s*刀', True),
        (r'(\d+\.?\d*)\s*元\s*(一|/)?\s*(刀|百万)', True),
        (r'(\d+\.?\d*)\s*(一|/)?\s*刀', True),
    ]
    for pat, use_ratio in patterns:
        m = re.search(pat, text)
        if m:
            val = float(m.group(1))
            disc = round(val / USD_RATIO, 2) if use_ratio else round(val, 2)
            remaining = text[m.end():].strip()
            purchaser = remaining if remaining and _ch_is_chinese_name(remaining) else None
            return disc, purchaser
    return _ch_parse_discount(text), None


def _ch_parse_channel_info(name, remark):
    default_purchaser = '奔云'
    default_supplier = 'XX'
    default_discount = 0.15

    if not name:
        name = ''

    if '@' not in name:
        supplier = default_supplier
        if remark and remark != 'None' and len(remark) > 0:
            supplier = remark.strip()
        return default_purchaser, supplier, default_discount, default_discount

    parts = [p.strip() for p in name.split('@') if p.strip()]
    purchaser = default_purchaser
    supplier = default_supplier
    discount = None
    pre_extracted_purchaser = None

    discount_idx = -1
    for i, part in enumerate(parts):
        disc, pot_purch = _ch_parse_discount_and_purchaser(part)
        if disc is not None:
            discount_idx = i
            discount = disc
            if pot_purch:
                pre_extracted_purchaser = pot_purch

    if discount is None:
        discount = default_discount
    discount_orig = discount

    has_discount = discount_idx >= 0
    parts_count = len(parts)

    def can_be_supplier(part, exclude_idx):
        if not part:
            return False
        idx = parts.index(part) if part in parts else -1
        if idx == exclude_idx or idx == discount_idx:
            return False
        if _ch_parse_discount(part) is not None:
            return False
        if re.match(r'^\d+\.\d+\.\d+\.\d+', part) or re.match(r'^[\d.]+:\d+', part):
            return False
        if re.match(r'^\d', part):
            return False
        if any(part.lower().startswith(m) or m in part.lower() for m in MODEL_NAMES):
            return False
        return True

    def can_be_purchaser(part, exclude_idx):
        if not part:
            return False
        idx = parts.index(part) if part in parts else -1
        if idx == exclude_idx or idx == discount_idx:
            return False
        if re.match(r'^\d', part):
            return False
        return _ch_is_purchaser_candidate(part) or _ch_is_chinese_name(part)

    if has_discount and parts_count > 4:
        if pre_extracted_purchaser:
            purchaser = pre_extracted_purchaser
        for si in [1, 2]:
            if si >= parts_count or si == discount_idx:
                continue
            if can_be_supplier(parts[si], -1):
                supplier = _ch_clean_supplier_name(parts[si])
                if supplier != 'XX':
                    break
        if not pre_extracted_purchaser:
            supplier_idx = parts.index(supplier) if supplier in parts else -1
            for i in range(parts_count - 1, -1, -1):
                if i == discount_idx or i == supplier_idx or i == 0:
                    continue
                if can_be_purchaser(parts[i], supplier_idx):
                    purchaser = parts[i]
                    break
    else:
        if pre_extracted_purchaser:
            purchaser = pre_extracted_purchaser
            for i in range(discount_idx - 1, 0, -1):
                if i == discount_idx or i == 0:
                    continue
                if can_be_purchaser(parts[i], -1):
                    purchaser = parts[i]
                    break
        else:
            for i in range(parts_count - 2, -1, -1):
                if i == discount_idx or i == 0:
                    continue
                if can_be_purchaser(parts[i], -1):
                    purchaser = parts[i]
                    break
        purchaser_idx = parts.index(purchaser) if purchaser in parts else -1
        for si in [1, 2]:
            if si >= parts_count or si == discount_idx or si == purchaser_idx:
                continue
            if can_be_supplier(parts[si], purchaser_idx):
                supplier = _ch_clean_supplier_name(parts[si])
                if supplier != 'XX':
                    break

    if supplier == default_supplier:
        for part in parts:
            if re.match(r'https?://', part) or re.match(r'^[\w.-]+\.[\w]+', part):
                idx = parts.index(part)
                p_idx = parts.index(purchaser) if purchaser in parts else -1
                if idx != discount_idx and idx != p_idx:
                    supplier = _ch_clean_supplier_name(part)
                    break

    purchaser = re.sub(r'[\/!@#$%^&]', '', purchaser)
    supplier = re.sub(r'[\/!@#$%^&]', '', supplier)
    return purchaser, supplier, discount_orig, discount


# ─── Token parsing ──────────────────────────────────────────────────

def _tk_parse_discount(text):
    if not text:
        return None
    text = text.strip()

    m = re.search(r'倍率[^0-9]*(\d+\.?\d*)', text)
    if m:
        return round(float(m.group(1)) * RATE_BASE, 6)

    m = re.search(r'(\d+\.?\d*)\s*倍', text)
    if m:
        return round(float(m.group(1)) * RATE_BASE, 6)

    m = re.search(r'(\d+\.?\d*)\s*%', text)
    if m:
        return round(float(m.group(1)) / 100, 2)

    m = re.search(r'(\d+\.?\d*)\s*折', text)
    if m:
        return round(float(m.group(1)) / 10, 2)

    m = re.search(r'(\d+\.?\d*)\s*￥', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*元\s*(一|/)?\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*(一|/)?\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'^(\d+\.?\d*)$', text)
    if m:
        val = float(m.group(1))
        if val < 10:
            return round(val / USD_RATIO, 2)
        elif val < 100:
            return round(val / 100, 2)

    m = re.search(r'(\d+\.?\d*)\s*分成', text)
    if m:
        val = float(m.group(1))
        if val <= 100:
            return round(val / 100, 2)

    m = re.search(r'(\d+\.?\d*)\s*%\s*分成', text)
    if m:
        return round(float(m.group(1)) / 100, 2)

    return None


def _tk_parse_token_discount(name):
    if not name or name == 'None':
        return 1.0
    discount = _tk_parse_discount(name.strip())
    return discount if discount is not None else 1.0


# ─── User parsing ───────────────────────────────────────────────────

def _us_parse_discount(text):
    if not text:
        return None
    text = text.strip()

    m = re.search(r'(\d+\.?\d*)\s*倍', text)
    if m:
        return round(float(m.group(1)), 2)

    m = re.search(r'(\d+\.?\d*)\s*%', text)
    if m:
        return round(float(m.group(1)) / 100, 2)

    if re.search(r'元\s*/\s*(张|次)', text):
        return 1.0

    m = re.search(r'(\d+\.?\d*)\s*折', text)
    if m:
        return round(float(m.group(1)) / 10, 2)

    m = re.search(r'(\d+\.?\d*)\s*￥', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*元\s*(一|/)?\s*(刀|百万)', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'^(\d+\.?\d*)\s*(一|/)?\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'(\d+\.?\d*)\s*(一|/)\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    m = re.search(r'^一\s*元\s*一', text)
    if m:
        return round(1.0 / USD_RATIO, 2)

    m = re.search(r'^(\d+\.?\d*)$', text)
    if m:
        val = float(m.group(1))
        if val < 10:
            return round(val / USD_RATIO, 2)
        elif val < 100:
            return round(val / 100, 2)

    m = re.search(r'(\d+\.?\d*)\s*元\s*(\d+\.?\d*)\s*刀', text)
    if m:
        return round(float(m.group(1)) / USD_RATIO, 2)

    return None


def _us_is_chinese_name(text):
    if not text or len(text) < 2 or len(text) > 4:
        return False
    if not all('一' <= ch <= '鿿' for ch in text):
        return False
    discount_kw = ['￥', '折', '刀', '元', '百万', '一刀', '一折', '/刀', '/元', '一']
    if any(kw in text for kw in discount_kw):
        return False
    if any(s in text for s in ['波转', '顺']):
        return False
    if any(w in text for w in ['备注', '额外', '信息', '说明', '测试', '临时', '代理', '直连']):
        return False
    return True


def _us_parse_user_discount_info(remark):
    default_sales = '奔云'
    default_discount = 0.05

    if not remark or remark == 'None':
        return default_sales, default_discount
    remark = remark.strip()

    all_chinese = all('一' <= ch <= '鿿' for ch in remark)
    if all_chinese and 1 <= len(remark) <= 4:
        non_name = ['备注', '额外', '信息', '说明', '测试', '临时', '代理', '直连', '官方', '稳定']
        if not any(w in remark for w in non_name):
            return remark, default_discount

    discount = None
    discount_start = 0
    discount_end = len(remark)

    patterns = [
        (r'(\d+\.?\d*)\s*元\s*(一|/)?\s*(刀|百万)', lambda m: round(float(m.group(1)) / USD_RATIO, 2)),
        (r'(\d+\.?\d*)\s*(一|/)?\s*刀', lambda m: round(float(m.group(1)) / USD_RATIO, 2)),
        (r'(\d+\.?\d*)\s*折', lambda m: round(float(m.group(1)) / 10, 2)),
        (r'(\d+\.?\d*)\s*元', lambda m: round(float(m.group(1)) / USD_RATIO, 2)),
        (r'(\d+\.?\d*)\s*￥', lambda m: round(float(m.group(1)) / USD_RATIO, 2)),
        (r'(\d+\.?\d*)\s*%', lambda m: round(float(m.group(1)) / 100, 2)),
        (r'(\d+\.?\d*)\s*倍', lambda m: round(float(m.group(1)), 2)),
    ]
    for pat, calc in patterns:
        m = re.search(pat, remark)
        if m and discount is None:
            discount = calc(m)
            discount_start = m.start()
            discount_end = m.end()

    if discount is None:
        m = re.search(r'(0\.\d+)', remark)
        if m and float(m.group(1)) < 1:
            discount = round(float(m.group(1)), 2)
            discount_start = m.start()
            discount_end = m.end()

    if discount is None:
        discount = default_discount

    prefix = remark[:discount_start]
    suffix = remark[discount_end:]

    for model in MODEL_NAMES:
        prefix = re.sub(re.escape(model), '', prefix, flags=re.IGNORECASE)
        suffix = re.sub(re.escape(model), '', suffix, flags=re.IGNORECASE)

    non_name = ['备注', '额外', '信息', '说明', '测试', '临时', '代理', '直连', '官方', '稳定',
                '专用', '分成', '折扣', '渠道', '倍率', '官转', '一', '折']

    def extract_chinese(text):
        m = re.search(r'([一-鿿]{1,4})', text)
        if m and not any(w in m.group(1) for w in non_name):
            return m.group(1)
        return None

    def clean_and_extract(text):
        cleaned = re.sub(r'[^一-鿿a-zA-Z]+', ' ', text.strip()).strip()
        if not cleaned:
            return None
        name = extract_chinese(cleaned)
        if name:
            return name
        m = re.search(r'([a-zA-Z]{2,6})', cleaned)
        if m:
            potential = m.group(1)
            exclude = ['pinova', 'csp', 'wzg', 'qn', 'ai'] + MODEL_NAMES
            if not any(potential.lower() == n.lower() for n in exclude):
                return potential
        return None

    sales_person = default_sales
    name = clean_and_extract(prefix)
    if name:
        sales_person = name
    else:
        name = clean_and_extract(suffix)
        if name:
            sales_person = name

    if sales_person == default_sales:
        for match in re.findall(r'[一-鿿]{1,4}', remark):
            if not any(w in match for w in non_name) and len(match) >= 1:
                sales_person = match
                break

    return sales_person, discount


# ─── Service functions ──────────────────────────────────────────────

_EXCEL_HEADERS = {
    'channels': ['渠道ID', '渠道名称', '采购员', '供应商', '原始折扣', '折扣'],
    'tokens': ['TokenID', 'UserID', 'name', 'discount'],
    'users': ['用户ID', '用户名称', '备注', '销售人员', '折扣'],
}

_EXCEL_WIDTHS = {
    'channels': [10, 40, 10, 12, 12, 10],
    'tokens': [10, 10, 40, 12],
    'users': [10, 20, 40, 10, 10],
}


def build_excel_bytes(table_type: str, headers: list[str], rows: list[tuple]) -> bytes:
    """Generate a proper .xlsx file using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ex_{table_type}"

    bold = Font(bold=True)
    widths = _EXCEL_WIDTHS.get(table_type, [])
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = bold
        if ci <= len(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = widths[ci - 1]

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

_EX_DB_COLS = {
    'channels': ['id', 'name', 'buyer', 'supplier', 'discount', 'discount_orig'],
    'tokens': ['id', 'discount'],
    'users': ['id', 'seller', 'remark', 'discount'],
}


async def parse_channels(site: str) -> dict:
    """Parse channels table → return Excel data (no DB write)."""
    config = AppConfig.load()
    db_name = config.db_name(site)

    rows = await db.fetch_all("SELECT id, name, remark FROM channels ORDER BY id", db=db_name)
    log.info(f"Parsing {len(rows)} channels for site {site}")

    parsed = []
    for r in rows:
        purchaser, supplier, discount_orig, discount = _ch_parse_channel_info(
            r.get('name') or '', r.get('remark') or ''
        )
        parsed.append((r['id'], r.get('name') or '', purchaser, supplier, discount, discount_orig))

    headers = _EXCEL_HEADERS['channels']
    return {
        "table": "ex_channels",
        "rows": len(parsed),
        "excel_headers": headers,
        "excel_data": parsed,
    }


async def parse_tokens(site: str) -> dict:
    """Parse tokens table → return Excel data (no DB write)."""
    config = AppConfig.load()
    db_name = config.db_name(site)

    rows = await db.fetch_all("SELECT id, user_id, name FROM tokens ORDER BY id", db=db_name)
    log.info(f"Parsing {len(rows)} tokens for site {site}")

    parsed = []
    for r in rows:
        discount = _tk_parse_token_discount(r.get('name') or '')
        parsed.append((r['id'], r.get('user_id') or 0, r.get('name') or '', discount))

    headers = _EXCEL_HEADERS['tokens']
    return {
        "table": "ex_tokens",
        "rows": len(parsed),
        "excel_headers": headers,
        "excel_data": parsed,
    }


async def parse_users(site: str) -> dict:
    """Parse users table → return Excel data (no DB write)."""
    config = AppConfig.load()
    db_name = config.db_name(site)

    rows = await db.fetch_all("SELECT id, username, remark FROM users ORDER BY id", db=db_name)
    log.info(f"Parsing {len(rows)} users for site {site}")

    parsed = []
    for r in rows:
        seller, discount = _us_parse_user_discount_info(r.get('remark') or '')
        parsed.append((r['id'], r.get('username') or '', r.get('remark') or '', seller, discount))

    headers = _EXCEL_HEADERS['users']
    return {
        "table": "ex_users",
        "rows": len(parsed),
        "excel_headers": headers,
        "excel_data": parsed,
    }
